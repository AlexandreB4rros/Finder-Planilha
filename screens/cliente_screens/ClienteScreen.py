from kivymd.uix.screen import MDScreen
from openpyxl.styles import PatternFill
from plyer import storagepath

from main import LiveApp

import os
import pickle
import httplib2
import google_auth_httplib2

from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.http import MediaFileUpload


class ClienteScreen(MDScreen):
    from kivy.properties import ObjectProperty

    text_field = ObjectProperty(None)

    def aplicar_config(self):
        self.Selecionar_Tema()

    def Selecionar_Tema(self):
        if self.ids.selecionar_tema_escuro.active:
            app = LiveApp()
            app.tema_escuro()

        elif self.ids.selecionar_tema_claro.active:
            app = LiveApp()
            app.tema_claro()
        else:
            pass

    def Limpar_Widget(self):
        self.ids.Nome_Cidade.text = ""
        self.ids.Nome_Pop.text = ""
        self.ids.Nome_Olt.text = "1"
        self.ids.Nome_Slot.text = ""
        self.ids.Nome_Pon.text = ""

    def Info_Fake(self):
        self.ids.Nome_Cidade.text = "XXXXXX"
        self.ids.Nome_Pop.text = "X"
        self.ids.Nome_Olt.text = "X"
        self.ids.Nome_Slot.text = "X"
        self.ids.Nome_Pon.text = "X"

    def proxima_tela_cto(self):
        from kivymd.uix.dialog import MDDialog

        campos = [self.ids.Nome_Cidade.text, self.ids.Nome_Pop.text, self.ids.Nome_Olt.text, self.ids.Nome_Slot.text,
                  self.ids.Nome_Pon.text
                  ]

        for campo in campos:
            if not campo:
                dialog = MDDialog(title='Atenção !',
                                  text='Todos os campos devem ser preenchidos...\n\n,DICA: Caso a CTO não esteja em conformidade com a documentação, preencha como "PILOTO1", "OLT:X", "SLOT:X","PON:X". Ao ser confirmado as informação pelo BKO, corrija as informações antes de salvar o levantamento.'
                                  )
                dialog.open()
                return
        self.manager.cidade_text = self.ids.Nome_Cidade.text
        self.manager.pop_text = self.ids.Nome_Pop.text
        self.manager.olt_text = self.ids.Nome_Olt.text
        self.manager.slot_text = self.ids.Nome_Slot.text
        self.manager.pon_text = self.ids.Nome_Pon.text

        self.manager.current = 'Selecionar_spliter'
        self.manager.transition.direction = 'right'

    def mudar_tela(self):
        self.manager.current = 'HomeScreen'
        self.manager.transition.direction = 'left'


class Selecionar_spliter(MDScreen):
    def Limpar_Widget(self):
        self.ids.cto.text = ""
        self.ids.obs.text = ""

    def Info_Fake(self):
        self.ids.cto.text = f"BBB-XXXX"

    def show_selected_box(self):
        if self.ids.umpraoito.active:
            self.vlor = '1/8'
            self.manager.current = 'tela1x8'
            self.manager.transition.direction = 'right'
        elif self.ids.umpradezeseis.active:
            self.vlor = '1/16'
            self.manager.current = 'tela1x16'
            self.manager.transition.direction = 'right'
        else:
            pass

    def botao_pressionado(self):
        self.manager.cto_text = self.ids.cto.text
        self.manager.obs_text = self.ids.obs.text
        if not hasattr(self.manager, 'obs_text'):
            self.manager.obs_text = " "

    def tela_anterior(self):
        self.manager.current = 'ClienteScreen'
        self.manager.transition.direction = 'left'

    def proxima_tela(self):
        from kivymd.uix.dialog import MDDialog

        campos = [self.ids.cto]
        for campo in campos:
            if not campo.text:
                dialog = MDDialog(title='Erro', text='Informe a CTO, indentifique o divisor para prosseguir...\n\nDICA: caso a cto não esteja em conformidade com a documentação, identifique a OLT seguido de letras ex: "AMN-XXXX".')
                dialog.open()
                return

        self.show_selected_box()
        self.botao_pressionado()

    def tela_home(self):
        self.manager.transition.direction = 'left'
        self.manager.current = 'HomeScreen'

    def camera_captura(self):
        self.manager.current = 'camera2'
        #self.manager.current.camera2


class tela1x8(MDScreen):
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build
    from google.auth.transport.requests import Request
    import os
    import pickle

    from kivy.properties import StringProperty, BooleanProperty
    from android.permissions import request_permissions, Permission  # type: ignore

    request_permissions([Permission.READ_EXTERNAL_STORAGE, Permission.WRITE_EXTERNAL_STORAGE])  # type: ignore
    documents_path = storagepath.get_documents_dir()

    def get_gdrive_service(self):
        SCOPES = ['https://www.googleapis.com/auth/drive.file']
        creds = None

        if os.path.exists('token.pickle'):
            with open('token.pickle', 'rb') as token:
                creds = pickle.load(token)
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(
                    'FinderPlanilhaCredenciaisDrive.json', SCOPES)
                creds = flow.run_local_server(port=0)
            with open('token.pickle', 'wb') as token:
                pickle.dump(creds, token)
        return build('drive', 'v3', credentials=creds)

    def upload_file(self, file_path, file_name):

        service = self.get_gdrive_service()
        file_metadata = {'name': file_name}
        media = MediaFileUpload(file_path, mimetype='application/octet-stream')
        file = service.files().create(body=file_metadata, media_body=media, fields='id').execute()
        print(f"File ID: {file.get('id')}")

    time_text01 = StringProperty(" ")
    time_text02 = StringProperty(" ")
    time_text03 = StringProperty(" ")
    time_text04 = StringProperty(" ")
    time_text05 = StringProperty(" ")
    time_text06 = StringProperty(" ")
    time_text07 = StringProperty(" ")
    time_text08 = StringProperty(" ")

    start_time01 = StringProperty("")
    start_time02 = StringProperty("")
    start_time03 = StringProperty("")
    start_time04 = StringProperty("")
    start_time05 = StringProperty("")
    start_time06 = StringProperty("")
    start_time07 = StringProperty("")
    start_time08 = StringProperty("")

    end_time01 = StringProperty("")
    end_time02 = StringProperty("")
    end_time03 = StringProperty("")
    end_time04 = StringProperty("")
    end_time05 = StringProperty("")
    end_time06 = StringProperty("")
    end_time07 = StringProperty("")
    end_time08 = StringProperty("")

    clock_clicked01 = BooleanProperty(False)
    clock_clicked02 = BooleanProperty(False)
    clock_clicked03 = BooleanProperty(False)
    clock_clicked04 = BooleanProperty(False)
    clock_clicked05 = BooleanProperty(False)
    clock_clicked06 = BooleanProperty(False)
    clock_clicked07 = BooleanProperty(False)
    clock_clicked08 = BooleanProperty(False)

    def Limpar_Hora(self):
        self.time_text01 = ""
        self.time_text02 = ""
        self.time_text03 = ""
        self.time_text04 = ""
        self.time_text05 = ""
        self.time_text06 = ""
        self.time_text07 = ""
        self.time_text08 = ""

        self.start_time01 = ""
        self.start_time02 = ""
        self.start_time03 = ""
        self.start_time04 = ""
        self.start_time05 = ""
        self.start_time06 = ""
        self.start_time07 = ""
        self.start_time08 = ""

        self.end_time01 = ""
        self.end_time02 = ""
        self.end_time03 = ""
        self.end_time04 = ""
        self.end_time05 = ""
        self.end_time06 = ""
        self.end_time07 = ""
        self.end_time08 = ""

        self.clock_clicked01 = ""
        self.clock_clicked02 = ""
        self.clock_clicked03 = ""
        self.clock_clicked04 = ""
        self.clock_clicked05 = ""
        self.clock_clicked06 = ""
        self.clock_clicked07 = ""
        self.clock_clicked08 = ""

        self.manager.StsP01_text = "*"
        self.manager.StsP02_text = "*"
        self.manager.StsP03_text = "*"
        self.manager.StsP04_text = "*"
        self.manager.StsP05_text = "*"
        self.manager.StsP06_text = "*"
        self.manager.StsP07_text = "*"
        self.manager.StsP08_text = "*"

        self.manager.SNP01_text = "-"
        self.manager.SNP02_text = "-"
        self.manager.SNP03_text = "-"
        self.manager.SNP04_text = "-"
        self.manager.SNP05_text = "-"
        self.manager.SNP06_text = "-"
        self.manager.SNP07_text = "-"
        self.manager.SNP08_text = "-"

        self.manager.PPP01_text = "-"
        self.manager.PPP02_text = "-"
        self.manager.PPP03_text = "-"
        self.manager.PPP04_text = "-"
        self.manager.PPP05_text = "-"
        self.manager.PPP06_text = "-"
        self.manager.PPP07_text = "-"
        self.manager.PPP08_text = "-"

        self.manager.Pwd01_text = "-"
        self.manager.Pwd02_text = "-"
        self.manager.Pwd03_text = "-"
        self.manager.Pwd04_text = "-"
        self.manager.Pwd05_text = "-"
        self.manager.Pwd06_text = "-"
        self.manager.Pwd07_text = "-"
        self.manager.Pwd08_text = "-"

        self.manager.Loid01_text = "-"
        self.manager.Loid02_text = "-"
        self.manager.Loid03_text = "-"
        self.manager.Loid04_text = "-"
        self.manager.Loid05_text = "-"
        self.manager.Loid06_text = "-"
        self.manager.Loid07_text = "-"
        self.manager.Loid08_text = "-"

        self.manager.Obs01_text = ""
        self.manager.Obs02_text = ""
        self.manager.Obs03_text = ""
        self.manager.Obs04_text = ""
        self.manager.Obs05_text = ""
        self.manager.Obs06_text = ""
        self.manager.Obs07_text = ""
        self.manager.Obs08_text = ""
    def fachina(self):
        self.Limpar_widget_1x8()
        self.Limpar_Hora()

    def save_timeP01(self):
        import time

        if not self.clock_clicked01:
            self.start_time01 = time.strftime("%H:%M")
            self.time_text01 = f"{self.start_time01}"
            self.clock_clicked01 = True
            print("Hora da Desconexão")
        else:
            self.end_time01 = time.strftime("%H:%M")
            self.time_text01 = f"{self.start_time01}/{self.end_time01}"
            self.clock_clicked01 = False
            print("Hora da Conexão / Desconexão:")
        print(self.time_text01)

    def save_timeP02(self):
        import time

        if not self.clock_clicked02:
            self.start_time02 = time.strftime("%H:%M")
            self.time_text02 = f"{self.start_time02}"
            self.clock_clicked02 = True
            print("Hora da Desconexão")
        else:
            self.end_time02 = time.strftime("%H:%M")
            self.time_text02 = f"{self.start_time02}/{self.end_time02}"
            self.clock_clicked02 = False
            print("Hora da Conexão / Desconexão:")
        print(self.time_text02)

    def save_timeP03(self):
        import time

        if not self.clock_clicked03:
            self.start_time03 = time.strftime("%H:%M")
            self.time_text03 = f"{self.start_time03}"
            self.clock_clicked03 = True
            print("Hora da Desconexão")
        else:
            self.end_time03 = time.strftime("%H:%M")
            self.time_text03 = f"{self.start_time03}/{self.end_time03}"
            self.clock_clicked03 = False
            print("Hora da Conexão / Desconexão:")
        print(self.time_text03)

    def save_timeP04(self):
        import time

        if not self.clock_clicked04:
            self.start_time04 = time.strftime("%H:%M")
            self.time_text04 = f"{self.start_time04}"
            self.clock_clicked04 = True
            print("Hora da Desconexão")
        else:
            self.end_time04 = time.strftime("%H:%M")
            self.time_text04 = f"{self.start_time04}/{self.end_time04}"
            self.clock_clicked04 = False
            print("Hora da Conexão / Desconexão:")
        print(self.time_text04)

    def save_timeP05(self):
        import time

        if not self.clock_clicked05:
            self.start_time05 = time.strftime("%H:%M")
            self.time_text05 = f"{self.start_time05}"
            self.clock_clicked05 = True
            print("Hora da Desconexão")
        else:
            self.end_time05 = time.strftime("%H:%M")
            self.time_text05 = f"{self.start_time05}/{self.end_time05}"
            self.clock_clicked05 = False
            print("Hora da Conexão / Desconexão:")
        print(self.time_text05)

    def save_timeP06(self):
        import time

        if not self.clock_clicked06:
            self.start_time06 = time.strftime("%H:%M")
            self.time_text06 = f"{self.start_time06}"
            self.clock_clicked06 = True
            print("Hora da Desconexão")
        else:
            self.end_time06 = time.strftime("%H:%M")
            self.time_text06 = f"{self.start_time06}/{self.end_time06}"
            self.clock_clicked06 = False
            print("Hora da Conexão / Desconexão:")
        print(self.time_text06)

    def save_timeP07(self):
        import time

        if not self.clock_clicked07:
            self.start_time07 = time.strftime("%H:%M")
            self.time_text07 = f"{self.start_time07}"
            self.clock_clicked07 = True
            print("Hora da Desconexão")
        else:
            self.end_time07 = time.strftime("%H:%M")
            self.time_text07 = f"{self.start_time07}/{self.end_time07}"
            self.clock_clicked07 = False
            print("Hora da Conexão / Desconexão:")
        print(self.time_text07)

    def save_timeP08(self):
        import time

        if not self.clock_clicked08:
            self.start_time08 = time.strftime("%H:%M")
            self.time_text08 = f"{self.start_time08}"
            self.clock_clicked08 = True
            print("Hora da Desconexão")
        else:
            self.end_time08 = time.strftime("%H:%M")
            self.time_text08 = f"{self.start_time08}/{self.end_time08}"
            self.clock_clicked08 = False
            print("Hora da Conexão / Desconexão:")
        print(self.time_text08)

    def salvar_cto(self):
        self.caixa_dialogo()

    def Copiar_Cto(self, instance):
        #ATENÇÃO - por medidas de segurança, não apague o instance

        from kivy.core.clipboard import Clipboard

        time01 = self.time_text01
        time02 = self.time_text02
        time03 = self.time_text03
        time04 = self.time_text04
        time05 = self.time_text05
        time06 = self.time_text06
        time07 = self.time_text07
        time08 = self.time_text08

        SNP01 = self.manager.SNP01_text.upper()
        SNP02 = self.manager.SNP02_text.upper()
        SNP03 = self.manager.SNP03_text.upper()
        SNP04 = self.manager.SNP04_text.upper()
        SNP05 = self.manager.SNP05_text.upper()
        SNP06 = self.manager.SNP06_text.upper()
        SNP07 = self.manager.SNP07_text.upper()
        SNP08 = self.manager.SNP08_text.upper()

        STSP01 = self.manager.StsP01_text.upper()
        STSP02 = self.manager.StsP02_text.upper()
        STSP03 = self.manager.StsP03_text.upper()
        STSP04 = self.manager.StsP04_text.upper()
        STSP05 = self.manager.StsP05_text.upper()
        STSP06 = self.manager.StsP06_text.upper()
        STSP07 = self.manager.StsP07_text.upper()
        STSP08 = self.manager.StsP08_text.upper()

        cidade_text = self.manager.cidade_text.upper()
        pop_text = self.manager.pop_text.upper()
        slot_txt = self.manager.slot_text.upper()
        pon_text = self.manager.pon_text.upper()
        OBS_CTO = self.manager.obs_text

        if not hasattr(self.manager, 'OBS_CTO'):
            self.manager.OBS_CTO = ""

        import re

        cto_text = self.manager.cto_text.upper()
        cto_text = re.sub(r"\s+", "-", cto_text)

        A01 = f"{cidade_text}-{pop_text}-SLOT:{slot_txt}-PON:{pon_text}-{cto_text} | 04-JOSÉ\n"
        A02 = f"P1 - {STSP01} - {time01} - {SNP01}"
        A03 = f"P2 - {STSP02} - {time02} - {SNP02}"
        A04 = f"P3 - {STSP03} - {time03} - {SNP03}"
        A05 = f"P4 - {STSP04} - {time04} - {SNP04}"
        A06 = f"P5 - {STSP05} - {time05} - {SNP05}"
        A07 = f"P6 - {STSP06} - {time06} - {SNP06}"
        A08 = f"P7 - {STSP07} - {time07} - {SNP07}"
        A09 = f"P8 - {STSP08} - {time08} - {SNP08}"
        A10 = f"\n{OBS_CTO}"

        all_portas = '\n'.join([A01, A02, A03, A04, A05, A06, A07, A08, A09, A10])
        Clipboard.copy(all_portas)

    def salvar_progresso(self, instance):
        #ATENÇÃO - por medidas de segurança, não apague o instance

        from kivy.core.clipboard import Clipboard

        time01 = self.time_text01
        time02 = self.time_text02
        time03 = self.time_text03
        time04 = self.time_text04
        time05 = self.time_text05
        time06 = self.time_text06
        time07 = self.time_text07
        time08 = self.time_text08

        SNP01 = self.manager.SNP01_text.upper()
        SNP02 = self.manager.SNP02_text.upper()
        SNP03 = self.manager.SNP03_text.upper()
        SNP04 = self.manager.SNP04_text.upper()
        SNP05 = self.manager.SNP05_text.upper()
        SNP06 = self.manager.SNP06_text.upper()
        SNP07 = self.manager.SNP07_text.upper()
        SNP08 = self.manager.SNP08_text.upper()

        STSP01 = self.manager.StsP01_text.upper()
        STSP02 = self.manager.StsP02_text.upper()
        STSP03 = self.manager.StsP03_text.upper()
        STSP04 = self.manager.StsP04_text.upper()
        STSP05 = self.manager.StsP05_text.upper()
        STSP06 = self.manager.StsP06_text.upper()
        STSP07 = self.manager.StsP07_text.upper()
        STSP08 = self.manager.StsP08_text.upper()

        cidade_text = self.manager.cidade_text.upper()
        pop_text = self.manager.pop_text.upper()
        slot_txt = self.manager.slot_text.upper()
        pon_text = self.manager.pon_text.upper()
        OBS_CTO = self.manager.obs_text

        if not hasattr(self.manager, 'OBS_CTO'):
            self.manager.OBS_CTO = ""

        import re

        cto_text = self.manager.cto_text.upper()
        cto_text = re.sub(r"\s+", "-", cto_text)

        A01 = f"{cidade_text}-{pop_text}-SLOT:{slot_txt}-PON:{pon_text}-{cto_text} | 04-JOSÉ\n"
        A02 = f"P1 - {STSP01} - {time01} - {SNP01}"
        A03 = f"P2 - {STSP02} - {time02} - {SNP02}"
        A04 = f"P3 - {STSP03} - {time03} - {SNP03}"
        A05 = f"P4 - {STSP04} - {time04} - {SNP04}"
        A06 = f"P5 - {STSP05} - {time05} - {SNP05}"
        A07 = f"P6 - {STSP06} - {time06} - {SNP06}"
        A08 = f"P7 - {STSP07} - {time07} - {SNP07}"
        A09 = f"P8 - {STSP08} - {time08} - {SNP08}"
        A10 = f"\n{OBS_CTO}"

        all_portas = '\n'.join([A01, A02, A03, A04, A05, A06, A07, A08, A09, A10])
        Clipboard.copy(all_portas)

        self.manager.current = 'Selecionar_spliter'
        self.manager.transition.direction = 'left'
        self.Upload_Drive()
        self.save_to_file2()

    def caixa_dialogo(self):
        import re

        from kivymd.uix.dialog import MDDialog
        from kivymd.uix.button import MDFlatButton

        if not hasattr(self.manager, 'SNP01_text'):
            self.manager.SNP01_text = "-"
        if not hasattr(self.manager, 'SNP02_text'):
            self.manager.SNP02_text = "-"
        if not hasattr(self.manager, 'SNP03_text'):
            self.manager.SNP03_text = "-"
        if not hasattr(self.manager, 'SNP04_text'):
            self.manager.SNP04_text = "-"
        if not hasattr(self.manager, 'SNP05_text'):
            self.manager.SNP05_text = "-"
        if not hasattr(self.manager, 'SNP06_text'):
            self.manager.SNP06_text = "-"
        if not hasattr(self.manager, 'SNP07_text'):
            self.manager.SNP07_text = "-"
        if not hasattr(self.manager, 'SNP08_text'):
            self.manager.SNP08_text = "-"

        if not hasattr(self.manager, 'StsP01_text'):
            self.manager.StsP01_text = "*"
        if not hasattr(self.manager, 'StsP02_text'):
            self.manager.StsP02_text = "*"
        if not hasattr(self.manager, 'StsP03_text'):
            self.manager.StsP03_text = "*"
        if not hasattr(self.manager, 'StsP04_text'):
            self.manager.StsP04_text = "*"
        if not hasattr(self.manager, 'StsP05_text'):
            self.manager.StsP05_text = "*"
        if not hasattr(self.manager, 'StsP06_text'):
            self.manager.StsP06_text = "*"
        if not hasattr(self.manager, 'StsP07_text'):
            self.manager.StsP07_text = "*"
        if not hasattr(self.manager, 'StsP08_text'):
            self.manager.StsP08_text = "*"

        cto_text = self.manager.cto_text.upper()
        cto_text = re.sub(r"\s+", "-", cto_text)
        OBS_CTO = self.manager.obs_text
        if not hasattr(self.manager, 'OBS_CTO'):
            self.manager.OBS_CTO = " "

        time01 = self.time_text01
        time02 = self.time_text02
        time03 = self.time_text03
        time04 = self.time_text04
        time05 = self.time_text05
        time06 = self.time_text06
        time07 = self.time_text07
        time08 = self.time_text08

        SNP01 = self.manager.SNP01_text.upper()
        SNP02 = self.manager.SNP02_text.upper()
        SNP03 = self.manager.SNP03_text.upper()
        SNP04 = self.manager.SNP04_text.upper()
        SNP05 = self.manager.SNP05_text.upper()
        SNP06 = self.manager.SNP06_text.upper()
        SNP07 = self.manager.SNP07_text.upper()
        SNP08 = self.manager.SNP08_text.upper()

        STSP01 = self.manager.StsP01_text.upper()
        STSP02 = self.manager.StsP02_text.upper()
        STSP03 = self.manager.StsP03_text.upper()
        STSP04 = self.manager.StsP04_text.upper()
        STSP05 = self.manager.StsP05_text.upper()
        STSP06 = self.manager.StsP06_text.upper()
        STSP07 = self.manager.StsP07_text.upper()
        STSP08 = self.manager.StsP08_text.upper()

        cidade_text = self.manager.cidade_text.upper()
        pop_text = self.manager.pop_text.upper()
        slot_txt = self.manager.slot_text.upper()
        pon_text = self.manager.pon_text.upper()

        porta1 = f"P1-{STSP01}-{time01}-{SNP01}"
        porta2 = f"P2-{STSP02}-{time02}-{SNP02}"
        porta3 = f"P3-{STSP03}-{time03}-{SNP03}"
        porta4 = f"P4-{STSP04}-{time04}-{SNP04}"
        porta5 = f"P5-{STSP05}-{time05}-{SNP05}"
        porta6 = f"P6-{STSP06}-{time06}-{SNP06}"
        porta7 = f"P7-{STSP07}-{time07}-{SNP07}"
        porta8 = f"P8-{STSP08}-{time08}-{SNP08}"

        dialogo = MDDialog(title=f"{cidade_text}\n\n{pop_text}-SLOT:{slot_txt}-PON:{pon_text}-{cto_text}",
                           text=f"\n{porta1}\n{porta2}\n{porta3}\n{porta4}\n{porta5}\n{porta6}\n{porta7}\n{porta8}\n\n{OBS_CTO}",
                           buttons=[
                               MDFlatButton(
                                   text="Upload Drive",
                                   theme_text_color="Custom",
                                   text_color=self.theme_cls.primary_color,
                                   on_release=self.Upload_Drive
                               ), #lambda *args: self.caixa_de_dialogo.dismiss()
                               MDFlatButton(
                                   text="Copiar",
                                   theme_text_color="Custom",
                                   text_color=self.theme_cls.primary_color,
                                   on_release=self.Copiar_Cto
                               ),
                               MDFlatButton(
                                   text="Salvar",
                                   theme_text_color="Custom",
                                   text_color=self.theme_cls.primary_color,
                                   on_release=self.salvar_progresso
                               ),

                           ],
                           )

        self.caixa_de_dialogo = dialogo
        self.caixa_de_dialogo.open()

    def tela_anterior(self):
        self.manager.current = 'Selecionar_spliter'
        self.manager.transition.direction = 'left'

    def tela_home(self):
        self.manager.current = 'HomeScreen'
        self.manager.transition.direction = 'left'

    def save_to_file2(self):
        import os
        import re
        import time
        import pandas as pd
        from plyer import storagepath
        from openpyxl import load_workbook

        #from android.permissions import request_permissions, Permission  # type: ignore

        if not hasattr(self.manager, 'StsP01_text'):
            self.manager.StsP01_text = "-"
        if not hasattr(self.manager, 'StsP02_text'):
            self.manager.StsP02_text = "-"
        if not hasattr(self.manager, 'StsP03_text'):
            self.manager.StsP03_text = "-"
        if not hasattr(self.manager, 'StsP04_text'):
            self.manager.StsP04_text = "-"
        if not hasattr(self.manager, 'StsP05_text'):
            self.manager.StsP05_text = "-"
        if not hasattr(self.manager, 'StsP06_text'):
            self.manager.StsP06_text = "-"
        if not hasattr(self.manager, 'StsP07_text'):
            self.manager.StsP07_text = "-"
        if not hasattr(self.manager, 'StsP08_text'):
            self.manager.StsP08_text = "-"

        if not hasattr(self.manager, 'SNP01_text'):
            self.manager.SNP01_text = "-"
        if not hasattr(self.manager, 'SNP02_text'):
            self.manager.SNP02_text = "-"
        if not hasattr(self.manager, 'SNP03_text'):
            self.manager.SNP03_text = "-"
        if not hasattr(self.manager, 'SNP04_text'):
            self.manager.SNP04_text = "-"
        if not hasattr(self.manager, 'SNP05_text'):
            self.manager.SNP05_text = "-"
        if not hasattr(self.manager, 'SNP06_text'):
            self.manager.SNP06_text = "-"
        if not hasattr(self.manager, 'SNP07_text'):
            self.manager.SNP07_text = "-"
        if not hasattr(self.manager, 'SNP08_text'):
            self.manager.SNP08_text = "-"

        if not hasattr(self.manager, 'PPP01_text'):
            self.manager.PPP01_text = "-"
        if not hasattr(self.manager, 'PPP02_text'):
            self.manager.PPP02_text = "-"
        if not hasattr(self.manager, 'PPP03_text'):
            self.manager.PPP03_text = "-"
        if not hasattr(self.manager, 'PPP04_text'):
            self.manager.PPP04_text = "-"
        if not hasattr(self.manager, 'PPP05_text'):
            self.manager.PPP05_text = "-"
        if not hasattr(self.manager, 'PPP06_text'):
            self.manager.PPP06_text = "-"
        if not hasattr(self.manager, 'PPP07_text'):
            self.manager.PPP07_text = "-"
        if not hasattr(self.manager, 'PPP08_text'):
            self.manager.PPP08_text = "-"

        if not hasattr(self.manager, 'Pwd01_text'):
            self.manager.Pwd01_text = "-"
        if not hasattr(self.manager, 'Pwd02_text'):
            self.manager.Pwd02_text = "-"
        if not hasattr(self.manager, 'Pwd03_text'):
            self.manager.Pwd03_text = "-"
        if not hasattr(self.manager, 'Pwd04_text'):
            self.manager.Pwd04_text = "-"
        if not hasattr(self.manager, 'Pwd05_text'):
            self.manager.Pwd05_text = "-"
        if not hasattr(self.manager, 'Pwd06_text'):
            self.manager.Pwd06_text = "-"
        if not hasattr(self.manager, 'Pwd07_text'):
            self.manager.Pwd07_text = "-"
        if not hasattr(self.manager, 'Pwd08_text'):
            self.manager.Pwd08_text = "-"

        if not hasattr(self.manager, 'Loid01_text'):
            self.manager.Loid01_text = "-"
        if not hasattr(self.manager, 'Loid02_text'):
            self.manager.Loid02_text = "-"
        if not hasattr(self.manager, 'Loid03_text'):
            self.manager.Loid03_text = "-"
        if not hasattr(self.manager, 'Loid04_text'):
            self.manager.Loid04_text = "-"
        if not hasattr(self.manager, 'Loid05_text'):
            self.manager.Loid05_text = "-"
        if not hasattr(self.manager, 'Loid06_text'):
            self.manager.Loid06_text = "-"
        if not hasattr(self.manager, 'Loid07_text'):
            self.manager.Loid07_text = "-"
        if not hasattr(self.manager, 'Loid08_text'):
            self.manager.Loid08_text = "-"

        if not hasattr(self.manager, 'Obs01_text'):
            self.manager.Obs01_text = "-"
        if not hasattr(self.manager, 'Obs02_text'):
            self.manager.Obs02_text = "-"
        if not hasattr(self.manager, 'Obs03_text'):
            self.manager.Obs03_text = "-"
        if not hasattr(self.manager, 'Obs04_text'):
            self.manager.Obs04_text = "-"
        if not hasattr(self.manager, 'Obs05_text'):
            self.manager.Obs05_text = "-"
        if not hasattr(self.manager, 'Obs06_text'):
            self.manager.Obs06_text = "-"
        if not hasattr(self.manager, 'Obs07_text'):
            self.manager.Obs07_text = "-"
        if not hasattr(self.manager, 'Obs08_text'):
            self.manager.Obs08_text = "-"

        StsP01 = self.manager.StsP01_text.upper()
        StsP02 = self.manager.StsP02_text.upper()
        StsP03 = self.manager.StsP03_text.upper()
        StsP04 = self.manager.StsP04_text.upper()
        StsP05 = self.manager.StsP05_text.upper()
        StsP06 = self.manager.StsP06_text.upper()
        StsP07 = self.manager.StsP07_text.upper()
        StsP08 = self.manager.StsP08_text.upper()

        SNP01 = self.manager.SNP01_text.upper()
        SNP02 = self.manager.SNP02_text.upper()
        SNP03 = self.manager.SNP03_text.upper()
        SNP04 = self.manager.SNP04_text.upper()
        SNP05 = self.manager.SNP05_text.upper()
        SNP06 = self.manager.SNP06_text.upper()
        SNP07 = self.manager.SNP07_text.upper()
        SNP08 = self.manager.SNP08_text.upper()

        PPP01 = self.manager.PPP01_text.upper()
        PPP02 = self.manager.PPP02_text.upper()
        PPP03 = self.manager.PPP03_text.upper()
        PPP04 = self.manager.PPP04_text.upper()
        PPP05 = self.manager.PPP05_text.upper()
        PPP06 = self.manager.PPP06_text.upper()
        PPP07 = self.manager.PPP07_text.upper()
        PPP08 = self.manager.PPP08_text.upper()

        Pwd01 = self.manager.Pwd01_text.upper()
        Pwd02 = self.manager.Pwd02_text.upper()
        Pwd03 = self.manager.Pwd03_text.upper()
        Pwd04 = self.manager.Pwd04_text.upper()
        Pwd05 = self.manager.Pwd05_text.upper()
        Pwd06 = self.manager.Pwd06_text.upper()
        Pwd07 = self.manager.Pwd07_text.upper()
        Pwd08 = self.manager.Pwd08_text.upper()

        Loid01 = self.manager.Loid01_text.upper()
        Loid02 = self.manager.Loid02_text.upper()
        Loid03 = self.manager.Loid03_text.upper()
        Loid04 = self.manager.Loid04_text.upper()
        Loid05 = self.manager.Loid05_text.upper()
        Loid06 = self.manager.Loid06_text.upper()
        Loid07 = self.manager.Loid07_text.upper()
        Loid08 = self.manager.Loid08_text.upper()

        Obs01 = self.manager.Obs01_text.upper()
        Obs02 = self.manager.Obs02_text.upper()
        Obs03 = self.manager.Obs03_text.upper()
        Obs04 = self.manager.Obs04_text.upper()
        Obs05 = self.manager.Obs05_text.upper()
        Obs06 = self.manager.Obs06_text.upper()
        Obs07 = self.manager.Obs07_text.upper()
        Obs08 = self.manager.Obs08_text.upper()

        ST01 = self.start_time01
        ST02 = self.start_time02
        ST03 = self.start_time03
        ST04 = self.start_time04
        ST05 = self.start_time05
        ST06 = self.start_time06
        ST07 = self.start_time07
        ST08 = self.start_time08

        ET01 = self.end_time01
        ET02 = self.end_time02
        ET03 = self.end_time03
        ET04 = self.end_time04
        ET05 = self.end_time05
        ET06 = self.end_time06
        ET07 = self.end_time07
        ET08 = self.end_time08

        objetos = [StsP01, StsP02, StsP03, StsP04, StsP05, StsP06, StsP07, StsP08]

        contador_hashtag = 0
        contador_ok = 0

        for obj in objetos:
            contador_hashtag += obj.count("#")
            contador_ok += obj.count("OK")

        PVG = contador_hashtag + contador_ok

        cidade_text = self.manager.cidade_text.upper()
        pop_text = self.manager.pop_text.upper()
        slot_txt = self.manager.slot_text.upper()
        pon_text = self.manager.pon_text.upper()
        data_hoje = time.strftime("%d/%m/%Y")
        cto_text = self.manager.cto_text.upper()
        cto_text = re.sub(r"\s+", "-", cto_text)
        OBS_CTO = self.manager.obs_text

        if not hasattr(self.manager, 'OBS_CTO'):
            self.manager.OBS_CTO = " "

        id_tec = "04 - JOSÉ"
        slotpon = f'SLOT {slot_txt} - PON {pon_text}'
        slotpon2 = f'{slot_txt}/{pon_text}'

        #request_permissions([Permission.READ_EXTERNAL_STORAGE, Permission.WRITE_EXTERNAL_STORAGE])  # type: ignore
        #documents_path = storagepath.get_documents_dir()

        folder_path = os.path.join(self.documents_path, 'FinderPlanilha', f'{cidade_text}', f'{pop_text}', f'{slotpon}',
                                   'Planilha')

        #folder_path = os.getcwd()

        os.makedirs(folder_path, exist_ok=True)
        file_path = os.path.join(folder_path, f'{cidade_text}-{pop_text}-{slotpon}.xlsx')

        CTO_PLANILHA = {
            'DATA|CTO': [data_hoje, cto_text, cto_text, cto_text, cto_text, cto_text, cto_text, cto_text, cto_text,
                         " "],
            'EQUIPE|PORTA': [id_tec, 1, 2, 3, 4, 5, 6, 7, 8, " "],
            'CIDADE|STATUS': [cidade_text, StsP01, StsP02, StsP03, StsP04, StsP05, StsP06, StsP07, StsP08, " "],
            'POP|DESC.': [pop_text, ST01, ST02, ST03, ST04, ST05, ST06, ST07, ST08, " "],
            'SLOT|CONEX.': [slot_txt, ET01, ET02, ET03, ET04, ET05, ET06, ET07, ET08, " "],
            'PON|FSAN/SN': [pon_text, SNP01, SNP02, SNP03, SNP04, SNP05, SNP06, SNP07, SNP08, " "],
            'PPPoE': ["-", PPP01, PPP02, PPP03, PPP04, PPP05, PPP06, PPP07, PPP08, " "],
            'PWD': ["-", Pwd01, Pwd02, Pwd03, Pwd04, Pwd05, Pwd06, Pwd07, Pwd08, " "],
            'LOID': ["-", Loid01, Loid02, Loid03, Loid04, Loid05, Loid06, Loid07, Loid08, " "],
            'TOTAL CLIENTES|OBS.': [PVG, Obs01, Obs02, Obs03, Obs04, Obs05, Obs06, Obs07, Obs08 + OBS_CTO, " "],
        }

        CTO_RESUMO = {
            'SPLITAGEM': ["1/8"],
            'CTOs': [cto_text],
            'SLOT/PON': [slotpon2],
            'EQUIPE': [id_tec],
            'DROPs REDEX': [contador_ok],
            'DROPs ATIVOS': [" "],
            'DROPs N/I REDEX': [contador_hashtag],
            'TOTAL DROPs P/ CTO': [PVG],
            'TOTAL DROPs N/I': [' '],
            'DATA': [data_hoje],
        }

        df1 = pd.DataFrame(CTO_PLANILHA)
        df2 = pd.DataFrame(CTO_RESUMO)

        if os.path.exists(file_path):
            book = load_workbook(file_path)
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                writer.book = book
                writer.sheets = {ws.title: ws for ws in book.worksheets}
                if 'CTO_PLANILHA' in writer.sheets:
                    startrow1 = writer.sheets['CTO_PLANILHA'].max_row
                else:
                    startrow1 = 0

                if 'CTO_RESUMO' in writer.sheets:
                    startrow2 = writer.sheets['CTO_RESUMO'].max_row
                else:
                    startrow2 = 0

                df1.to_excel(writer, sheet_name='CTO_PLANILHA', startrow=startrow1, index=False, header=False)
                df2.to_excel(writer, sheet_name='CTO_RESUMO', startrow=startrow2, index=False, header=False)
        else:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df1.to_excel(writer, sheet_name='CTO_PLANILHA', index=False)
                df2.to_excel(writer, sheet_name='CTO_RESUMO', index=False)

        from openpyxl import load_workbook
        from openpyxl.styles import Border, Side

        # Carregue o arquivo XLSX existente
        book = load_workbook(file_path)

        # Selecione a planilha desejada (por exemplo, 'CTO_PLANILHA')
        sheet1 = book['CTO_PLANILHA']
        sheet1.column_dimensions['A'].width = 11
        sheet1.column_dimensions['B'].width = 15
        sheet1.column_dimensions['C'].width = 15
        sheet1.column_dimensions['D'].width = 12
        sheet1.column_dimensions['E'].width = 12
        sheet1.column_dimensions['F'].width = 20
        sheet1.column_dimensions['G'].width = 10
        sheet1.column_dimensions['H'].width = 10
        sheet1.column_dimensions['I'].width = 10
        sheet1.column_dimensions['J'].width = 20

        sheet2 = book['CTO_RESUMO']
        sheet2.column_dimensions['A'].width = 12
        sheet2.column_dimensions['B'].width = 17
        sheet2.column_dimensions['C'].width = 16
        sheet2.column_dimensions['D'].width = 13
        sheet2.column_dimensions['E'].width = 15
        sheet2.column_dimensions['F'].width = 18
        sheet2.column_dimensions['G'].width = 10
        sheet2.column_dimensions['H'].width = 10
        sheet2.column_dimensions['I'].width = 10
        sheet2.column_dimensions['J'].width = 25

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        sheet1['A1'].fill = PatternFill(fill_type='solid', start_color='5cb800')
        sheet1['B1'].fill = PatternFill(fill_type='solid', start_color='5cb800')
        sheet1['C1'].fill = PatternFill(fill_type='solid', start_color='5cb800')
        sheet1['D1'].fill = PatternFill(fill_type='solid', start_color='5cb800')
        sheet1['E1'].fill = PatternFill(fill_type='solid', start_color='5cb800')
        sheet1['F1'].fill = PatternFill(fill_type='solid', start_color='5cb800')
        sheet1['G1'].fill = PatternFill(fill_type='solid', start_color='5cb800')
        sheet1['H1'].fill = PatternFill(fill_type='solid', start_color='5cb800')
        sheet1['I1'].fill = PatternFill(fill_type='solid', start_color='5cb800')
        sheet1['J1'].fill = PatternFill(fill_type='solid', start_color='5cb800')

        # Percorra as células e aplique a borda
        for row in sheet1.iter_rows():
            for cell in row:
                cell.border = thin_border

        for row in sheet2.iter_rows():
            for cell in row:
                cell.border = thin_border

        # Salve o arquivo
        book.save(file_path)

        self.Limpar_widget_1x8()
        self.Limpar_Hora()

    def Upload_Drive(self, *args):
        import os

        cidade_text = self.manager.cidade_text.upper()
        pop_text = self.manager.pop_text.upper()
        slot_txt = self.manager.slot_text.upper()
        pon_text = self.manager.pon_text.upper()
        slotpon = f'SLOT {slot_txt} - PON {pon_text}'

        folder_path = os.path.join(self.documents_path, 'FinderPlanilha', f'{cidade_text}', f'{pop_text}', f'{slotpon}',
                                   'Planilha')

        folder_path1 = f'{folder_path}/{cidade_text}-{pop_text}-{slotpon}.xlsx'
        name_folder = f'{cidade_text}-{pop_text}-{slotpon}.xlsx'

        self.upload_file(f'{folder_path1}', f'{name_folder}')

    def Limpar_widget_1x8(self):
        self.manager.get_screen('PA01').clear_text()
        self.manager.get_screen('PA02').clear_text()
        self.manager.get_screen('PA03').clear_text()
        self.manager.get_screen('PA04').clear_text()
        self.manager.get_screen('PA05').clear_text()
        self.manager.get_screen('PA06').clear_text()
        self.manager.get_screen('PA07').clear_text()
        self.manager.get_screen('PA08').clear_text()


class tela1x16(MDScreen):
    from kivy.properties import StringProperty, BooleanProperty

    time_text01 = StringProperty(" ")
    time_text02 = StringProperty(" ")
    time_text03 = StringProperty(" ")
    time_text04 = StringProperty(" ")
    time_text05 = StringProperty(" ")
    time_text06 = StringProperty(" ")
    time_text07 = StringProperty(" ")
    time_text08 = StringProperty(" ")
    time_text09 = StringProperty(" ")
    time_text10 = StringProperty(" ")
    time_text11 = StringProperty(" ")
    time_text12 = StringProperty(" ")
    time_text13 = StringProperty(" ")
    time_text14 = StringProperty(" ")
    time_text15 = StringProperty(" ")
    time_text16 = StringProperty(" ")

    start_time01 = StringProperty("")
    start_time02 = StringProperty("")
    start_time03 = StringProperty("")
    start_time04 = StringProperty("")
    start_time05 = StringProperty("")
    start_time06 = StringProperty("")
    start_time07 = StringProperty("")
    start_time08 = StringProperty("")
    start_time09 = StringProperty("")
    start_time10 = StringProperty("")
    start_time11 = StringProperty("")
    start_time12 = StringProperty("")
    start_time13 = StringProperty("")
    start_time14 = StringProperty("")
    start_time15 = StringProperty("")
    start_time16 = StringProperty("")

    clock_clicked01 = BooleanProperty(False)
    clock_clicked02 = BooleanProperty(False)
    clock_clicked03 = BooleanProperty(False)
    clock_clicked04 = BooleanProperty(False)
    clock_clicked05 = BooleanProperty(False)
    clock_clicked06 = BooleanProperty(False)
    clock_clicked07 = BooleanProperty(False)
    clock_clicked08 = BooleanProperty(False)
    clock_clicked09 = BooleanProperty(False)
    clock_clicked10 = BooleanProperty(False)
    clock_clicked11 = BooleanProperty(False)
    clock_clicked12 = BooleanProperty(False)
    clock_clicked13 = BooleanProperty(False)
    clock_clicked14 = BooleanProperty(False)
    clock_clicked15 = BooleanProperty(False)
    clock_clicked16 = BooleanProperty(False)

    end_time01 = StringProperty("")
    end_time02 = StringProperty("")
    end_time03 = StringProperty("")
    end_time04 = StringProperty("")
    end_time05 = StringProperty("")
    end_time06 = StringProperty("")
    end_time07 = StringProperty("")
    end_time08 = StringProperty("")
    end_time09 = StringProperty("")
    end_time10 = StringProperty("")
    end_time11 = StringProperty("")
    end_time12 = StringProperty("")
    end_time13 = StringProperty("")
    end_time14 = StringProperty("")
    end_time15 = StringProperty("")
    end_time16 = StringProperty("")



    def Limpar_Hora(self):
        self.time_text01 = ""
        self.time_text02 = ""
        self.time_text03 = ""
        self.time_text04 = ""
        self.time_text05 = ""
        self.time_text06 = ""
        self.time_text07 = ""
        self.time_text08 = ""
        self.time_text09 = ""
        self.time_text10 = ""
        self.time_text11 = ""
        self.time_text12 = ""
        self.time_text13 = ""
        self.time_text14 = ""
        self.time_text15 = ""
        self.time_text16 = ""

        self.start_time01 = ""
        self.start_time02 = ""
        self.start_time03 = ""
        self.start_time04 = ""
        self.start_time05 = ""
        self.start_time06 = ""
        self.start_time07 = ""
        self.start_time08 = ""
        self.start_time09 = ""
        self.start_time10 = ""
        self.start_time11 = ""
        self.start_time12 = ""
        self.start_time13 = ""
        self.start_time14 = ""
        self.start_time15 = ""
        self.start_time16 = ""

        self.end_time01 = ""
        self.end_time02 = ""
        self.end_time03 = ""
        self.end_time04 = ""
        self.end_time05 = ""
        self.end_time06 = ""
        self.end_time07 = ""
        self.end_time08 = ""
        self.end_time09 = ""
        self.end_time10 = ""
        self.end_time11 = ""
        self.end_time12 = ""
        self.end_time13 = ""
        self.end_time14 = ""
        self.end_time15 = ""
        self.end_time16 = ""

        self.clock_clicked01 = ""
        self.clock_clicked02 = ""
        self.clock_clicked03 = ""
        self.clock_clicked04 = ""
        self.clock_clicked05 = ""
        self.clock_clicked06 = ""
        self.clock_clicked07 = ""
        self.clock_clicked08 = ""
        self.clock_clicked09 = ""
        self.clock_clicked10 = ""
        self.clock_clicked11 = ""
        self.clock_clicked12 = ""
        self.clock_clicked13 = ""
        self.clock_clicked14 = ""
        self.clock_clicked15 = ""
        self.clock_clicked16 = ""

        self.manager.StsP01_text = "*"
        self.manager.StsP02_text = "*"
        self.manager.StsP03_text = "*"
        self.manager.StsP04_text = "*"
        self.manager.StsP05_text = "*"
        self.manager.StsP06_text = "*"
        self.manager.StsP07_text = "*"
        self.manager.StsP08_text = "*"
        self.manager.StsP09_text = "*"
        self.manager.StsP10_text = "*"
        self.manager.StsP11_text = "*"
        self.manager.StsP12_text = "*"
        self.manager.StsP13_text = "*"
        self.manager.StsP14_text = "*"
        self.manager.StsP15_text = "*"
        self.manager.StsP16_text = "*"

        self.manager.SNP01_text = "-"
        self.manager.SNP02_text = "-"
        self.manager.SNP03_text = "-"
        self.manager.SNP04_text = "-"
        self.manager.SNP05_text = "-"
        self.manager.SNP06_text = "-"
        self.manager.SNP07_text = "-"
        self.manager.SNP08_text = "-"
        self.manager.SNP09_text = "-"
        self.manager.SNP10_text = "-"
        self.manager.SNP11_text = "-"
        self.manager.SNP12_text = "-"
        self.manager.SNP13_text = "-"
        self.manager.SNP14_text = "-"
        self.manager.SNP15_text = "-"
        self.manager.SNP16_text = "-"

        self.manager.PPP01_text = "-"
        self.manager.PPP02_text = "-"
        self.manager.PPP03_text = "-"
        self.manager.PPP04_text = "-"
        self.manager.PPP05_text = "-"
        self.manager.PPP06_text = "-"
        self.manager.PPP07_text = "-"
        self.manager.PPP08_text = "-"
        self.manager.PPP09_text = "-"
        self.manager.PPP10_text = "-"
        self.manager.PPP11_text = "-"
        self.manager.PPP12_text = "-"
        self.manager.PPP13_text = "-"
        self.manager.PPP14_text = "-"
        self.manager.PPP15_text = "-"
        self.manager.PPP16_text = "-"

        self.manager.Pwd01_text = "-"
        self.manager.Pwd02_text = "-"
        self.manager.Pwd03_text = "-"
        self.manager.Pwd04_text = "-"
        self.manager.Pwd05_text = "-"
        self.manager.Pwd06_text = "-"
        self.manager.Pwd07_text = "-"
        self.manager.Pwd08_text = "-"
        self.manager.Pwd09_text = "-"
        self.manager.Pwd10_text = "-"
        self.manager.Pwd11_text = "-"
        self.manager.Pwd12_text = "-"
        self.manager.Pwd13_text = "-"
        self.manager.Pwd14_text = "-"
        self.manager.Pwd15_text = "-"
        self.manager.Pwd16_text = "-"

        self.manager.Loid01_text = "-"
        self.manager.Loid02_text = "-"
        self.manager.Loid03_text = "-"
        self.manager.Loid04_text = "-"
        self.manager.Loid05_text = "-"
        self.manager.Loid06_text = "-"
        self.manager.Loid07_text = "-"
        self.manager.Loid08_text = "-"
        self.manager.Loid09_text = "-"
        self.manager.Loid10_text = "-"
        self.manager.Loid11_text = "-"
        self.manager.Loid12_text = "-"
        self.manager.Loid13_text = "-"
        self.manager.Loid14_text = "-"
        self.manager.Loid15_text = "-"
        self.manager.Loid16_text = "-"

        self.manager.Obs01_text = ""
        self.manager.Obs02_text = ""
        self.manager.Obs03_text = ""
        self.manager.Obs04_text = ""
        self.manager.Obs05_text = ""
        self.manager.Obs06_text = ""
        self.manager.Obs07_text = ""
        self.manager.Obs08_text = ""
        self.manager.Obs09_text = ""
        self.manager.Obs10_text = ""
        self.manager.Obs11_text = ""
        self.manager.Obs12_text = ""
        self.manager.Obs13_text = ""
        self.manager.Obs14_text = ""
        self.manager.Obs15_text = ""
        self.manager.Obs16_text = ""

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        self.StsP01 = ""  # Defina o atributo StsP01
    def fachina(self):
        self.Limpar_widget_1x16()
        self.Limpar_Hora()
    def save_timeP01(self):
        import time

        if not self.clock_clicked01:
            self.start_time01 = time.strftime("%H:%M")
            self.time_text01 = f"{self.start_time01}"
            self.clock_clicked01 = True
            print("Hora da Desconexão")
        else:
            self.end_time01 = time.strftime("%H:%M")
            self.time_text01 = f"{self.start_time01}/{self.end_time01}"
            self.clock_clicked01 = False
            print("Hora da Conexão / Desconexão:")
        print(self.time_text01)

    def save_timeP02(self):
        import time

        if not self.clock_clicked02:
            self.start_time02 = time.strftime("%H:%M")
            self.time_text02 = f"{self.start_time02}"
            self.clock_clicked02 = True
            print("Hora da Desconexão")
        else:
            self.end_time02 = time.strftime("%H:%M")
            self.time_text02 = f"{self.start_time02}/{self.end_time02}"
            self.clock_clicked02 = False
            print("Hora da Conexão / Desconexão:")
        print(self.time_text02)

    def save_timeP03(self):
        import time

        if not self.clock_clicked03:
            self.start_time03 = time.strftime("%H:%M")
            self.time_text03 = f"{self.start_time03}"
            self.clock_clicked03 = True
            print("Hora da Desconexão")
        else:
            self.end_time03 = time.strftime("%H:%M")
            self.time_text03 = f"{self.start_time03}/{self.end_time03}"
            self.clock_clicked03 = False
            print("Hora da Conexão / Desconexão:")
        print(self.time_text03)

    def save_timeP04(self):
        import time

        if not self.clock_clicked04:
            self.start_time04 = time.strftime("%H:%M")
            self.time_text04 = f"{self.start_time04}"
            self.clock_clicked04 = True
            print("Hora da Desconexão")
        else:
            self.end_time04 = time.strftime("%H:%M")
            self.time_text04 = f"{self.start_time04}/{self.end_time04}"
            self.clock_clicked04 = False
            print("Hora da Conexão / Desconexão:")
        print(self.time_text04)

    def save_timeP05(self):
        import time

        if not self.clock_clicked05:
            self.start_time05 = time.strftime("%H:%M")
            self.time_text05 = f"{self.start_time05}"
            self.clock_clicked05 = True
            print("Hora da Desconexão")
        else:
            self.end_time05 = time.strftime("%H:%M")
            self.time_text05 = f"{self.start_time05}/{self.end_time05}"
            self.clock_clicked05 = False
            print("Hora da Conexão / Desconexão:")
        print(self.time_text05)

    def save_timeP06(self):
        import time

        if not self.clock_clicked06:
            self.start_time06 = time.strftime("%H:%M")
            self.time_text06 = f"{self.start_time06}"
            self.clock_clicked06 = True
            print("Hora da Desconexão")
        else:
            self.end_time06 = time.strftime("%H:%M")
            self.time_text06 = f"{self.start_time06}/{self.end_time06}"
            self.clock_clicked06 = False
            print("Hora da Conexão / Desconexão:")
        print(self.time_text06)

    def save_timeP07(self):
        import time

        if not self.clock_clicked07:
            self.start_time07 = time.strftime("%H:%M")
            self.time_text07 = f"{self.start_time07}"
            self.clock_clicked07 = True
            print("Hora da Desconexão")
        else:
            self.end_time07 = time.strftime("%H:%M")
            self.time_text07 = f"{self.start_time07}/{self.end_time07}"
            self.clock_clicked07 = False
            print("Hora da Conexão / Desconexão:")
        print(self.time_text07)

    def save_timeP08(self):
        import time

        if not self.clock_clicked08:
            self.start_time08 = time.strftime("%H:%M")
            self.time_text08 = f"{self.start_time08}"
            self.clock_clicked08 = True
            print("Hora da Desconexão")
        else:
            self.end_time08 = time.strftime("%H:%M")
            self.time_text08 = f"{self.start_time08}/{self.end_time08}"
            self.clock_clicked08 = False
            print("Hora da Conexão / Desconexão:")
        print(self.time_text08)

    def save_timeP09(self):
        import time

        if not self.clock_clicked09:
            self.start_time09 = time.strftime("%H:%M")
            self.time_text09 = f"{self.start_time09}"
            self.clock_clicked09 = True
            print("Hora da Desconexão")
        else:
            self.end_time09 = time.strftime("%H:%M")
            self.time_text09 = f"{self.start_time09}/{self.end_time09}"
            self.clock_clicked09 = False
            print("Hora da Conexão / Desconexão:")
        print(self.time_text09)

    def save_timeP10(self):
        import time

        if not self.clock_clicked10:
            self.start_time10 = time.strftime("%H:%M")
            self.time_text10 = f"{self.start_time10}"
            self.clock_clicked10 = True
            print("Hora da Desconexão")
        else:
            self.end_time10 = time.strftime("%H:%M")
            self.time_text10 = f"{self.start_time10}/{self.end_time10}"
            self.clock_clicked10 = False
            print("Hora da Conexão / Desconexão:")
        print(self.time_text10)

    def save_timeP11(self):
        import time

        if not self.clock_clicked11:
            self.start_time11 = time.strftime("%H:%M")
            self.time_text11 = f"{self.start_time11}"
            self.clock_clicked11 = True
            print("Hora da Desconexão")
        else:
            self.end_time11 = time.strftime("%H:%M")
            self.time_text11 = f"{self.start_time11}/{self.end_time11}"
            self.clock_clicked11 = False
            print("Hora da Conexão / Desconexão:")
        print(self.time_text11)

    def save_timeP12(self):
        import time

        if not self.clock_clicked12:
            self.start_time12 = time.strftime("%H:%M")
            self.time_text12 = f"{self.start_time12}"
            self.clock_clicked12 = True
            print("Hora da Desconexão")
        else:
            self.end_time12 = time.strftime("%H:%M")
            self.time_text12 = f"{self.start_time12}/{self.end_time12}"
            self.clock_clicked12 = False
            print("Hora da Conexão / Desconexão:")
        print(self.time_text12)

    def save_timeP13(self):
        import time

        if not self.clock_clicked13:
            self.start_time13 = time.strftime("%H:%M")
            self.time_text13 = f"{self.start_time13}"
            self.clock_clicked13 = True
            print("Hora da Desconexão")
        else:
            self.end_time13 = time.strftime("%H:%M")
            self.time_text13 = f"{self.start_time13}/{self.end_time13}"
            self.clock_clicked13 = False
            print("Hora da Conexão / Desconexão:")
        print(self.time_text13)

    def save_timeP14(self):
        import time

        if not self.clock_clicked14:
            self.start_time14 = time.strftime("%H:%M")
            self.time_text14 = f"{self.start_time14}"
            self.clock_clicked14 = True
            print("Hora da Desconexão")
        else:
            self.end_time14 = time.strftime("%H:%M")
            self.time_text14 = f"{self.start_time14}/{self.end_time14}"
            self.clock_clicked14 = False
            print("Hora da Conexão / Desconexão:")
        print(self.time_text14)

    def save_timeP15(self):
        import time

        if not self.clock_clicked15:
            self.start_time15 = time.strftime("%H:%M")
            self.time_text15 = f"{self.start_time15}"
            self.clock_clicked15 = True
            print("Hora da Desconexão")
        else:
            self.end_time15 = time.strftime("%H:%M")
            self.time_text15 = f"{self.start_time15}/{self.end_time15}"
            self.clock_clicked15 = False
            print("Hora da Conexão / Desconexão:")
        print(self.time_text15)

    def save_timeP16(self):
        import time

        if not self.clock_clicked16:
            self.start_time16 = time.strftime("%H:%M")
            self.time_text16 = f"{self.start_time16}"
            self.clock_clicked16 = True
            print("Hora da Desconexão")
        else:
            self.end_time16 = time.strftime("%H:%M")
            self.time_text16 = f"{self.start_time16}/{self.end_time16}"
            self.clock_clicked16 = False
            print("Hora da Conexão / Desconexão:")
        print(self.time_text16)

    def salvar_cto(self):
        self.caixa_dialogo()


    def Copiar_Cto(self, instance):
        import re
        from kivy.core.clipboard import Clipboard

        time01 = self.time_text01
        time02 = self.time_text02
        time03 = self.time_text03
        time04 = self.time_text04
        time05 = self.time_text05
        time06 = self.time_text06
        time07 = self.time_text07
        time08 = self.time_text08
        time09 = self.time_text09
        time10 = self.time_text10
        time11 = self.time_text11
        time12 = self.time_text12
        time13 = self.time_text13
        time14 = self.time_text14
        time15 = self.time_text15
        time16 = self.time_text16

        SNP01 = self.manager.SNP01_text.upper()
        SNP02 = self.manager.SNP02_text.upper()
        SNP03 = self.manager.SNP03_text.upper()
        SNP04 = self.manager.SNP04_text.upper()
        SNP05 = self.manager.SNP05_text.upper()
        SNP06 = self.manager.SNP06_text.upper()
        SNP07 = self.manager.SNP07_text.upper()
        SNP08 = self.manager.SNP08_text.upper()
        SNP09 = self.manager.SNP09_text.upper()
        SNP10 = self.manager.SNP10_text.upper()
        SNP11 = self.manager.SNP11_text.upper()
        SNP12 = self.manager.SNP12_text.upper()
        SNP13 = self.manager.SNP13_text.upper()
        SNP14 = self.manager.SNP14_text.upper()
        SNP15 = self.manager.SNP15_text.upper()
        SNP16 = self.manager.SNP16_text.upper()

        STSP01 = self.manager.StsP01_text.upper()
        STSP02 = self.manager.StsP02_text.upper()
        STSP03 = self.manager.StsP03_text.upper()
        STSP04 = self.manager.StsP04_text.upper()
        STSP05 = self.manager.StsP05_text.upper()
        STSP06 = self.manager.StsP06_text.upper()
        STSP07 = self.manager.StsP07_text.upper()
        STSP08 = self.manager.StsP08_text.upper()
        STSP09 = self.manager.StsP09_text.upper()
        STSP10 = self.manager.StsP10_text.upper()
        STSP11 = self.manager.StsP11_text.upper()
        STSP12 = self.manager.StsP12_text.upper()
        STSP13 = self.manager.StsP13_text.upper()
        STSP14 = self.manager.StsP14_text.upper()
        STSP15 = self.manager.StsP15_text.upper()
        STSP16 = self.manager.StsP16_text.upper()

        cidade_text = self.manager.cidade_text.upper()
        pop_text = self.manager.pop_text.upper()
        slot_txt = self.manager.slot_text.upper()
        pon_text = self.manager.pon_text.upper()
        OBS_CTO = self.manager.obs_text

        if not hasattr(self.manager, 'OBS_CTO'):
            self.manager.OBS_CTO = " "

        cto_text = self.manager.cto_text.upper()
        cto_text = re.sub(r"\s+", "-", cto_text)

        A01 = f"{cidade_text}\n{pop_text}-SLOT:{slot_txt}-PON:{pon_text}-{cto_text} | 04-JOSÉ\n"
        A02 = f"P01 - {STSP01} - {time01} - {SNP01}"
        A03 = f"P02 - {STSP02} - {time02} - {SNP02}"
        A04 = f"P03 - {STSP03} - {time03} - {SNP03}"
        A05 = f"P04 - {STSP04} - {time04} - {SNP04}"
        A06 = f"P05 - {STSP05} - {time05} - {SNP05}"
        A07 = f"P06 - {STSP06} - {time06} - {SNP06}"
        A08 = f"P07 - {STSP07} - {time07} - {SNP07}"
        A09 = f"P08 - {STSP08} - {time08} - {SNP08}"
        A10 = f"P09 - {STSP09} - {time09} - {SNP09}"
        A11 = f"P10 - {STSP10} - {time10} - {SNP10}"
        A12 = f"P11 - {STSP11} - {time11} - {SNP11}"
        A13 = f"P12 - {STSP12} - {time12} - {SNP12}"
        A14 = f"P13 - {STSP13} - {time13} - {SNP13}"
        A15 = f"P14 - {STSP14} - {time14} - {SNP14}"
        A16 = f"P15 - {STSP15} - {time15} - {SNP15}"
        A17 = f"P16 - {STSP16} - {time16} - {SNP16}"
        A18 = f"\n{OBS_CTO}"

        all_portas = '\n'.join(
            [A01, A02, A03, A04, A05, A06, A07, A08, A09, A10, A11, A12, A13, A14, A15, A16, A17, A18])
        Clipboard.copy(all_portas)

    def salvar_progresso(self, instance):
        import re
        from kivy.core.clipboard import Clipboard

        time01 = self.time_text01
        time02 = self.time_text02
        time03 = self.time_text03
        time04 = self.time_text04
        time05 = self.time_text05
        time06 = self.time_text06
        time07 = self.time_text07
        time08 = self.time_text08
        time09 = self.time_text09
        time10 = self.time_text10
        time11 = self.time_text11
        time12 = self.time_text12
        time13 = self.time_text13
        time14 = self.time_text14
        time15 = self.time_text15
        time16 = self.time_text16

        SNP01 = self.manager.SNP01_text.upper()
        SNP02 = self.manager.SNP02_text.upper()
        SNP03 = self.manager.SNP03_text.upper()
        SNP04 = self.manager.SNP04_text.upper()
        SNP05 = self.manager.SNP05_text.upper()
        SNP06 = self.manager.SNP06_text.upper()
        SNP07 = self.manager.SNP07_text.upper()
        SNP08 = self.manager.SNP08_text.upper()
        SNP09 = self.manager.SNP09_text.upper()
        SNP10 = self.manager.SNP10_text.upper()
        SNP11 = self.manager.SNP11_text.upper()
        SNP12 = self.manager.SNP12_text.upper()
        SNP13 = self.manager.SNP13_text.upper()
        SNP14 = self.manager.SNP14_text.upper()
        SNP15 = self.manager.SNP15_text.upper()
        SNP16 = self.manager.SNP16_text.upper()

        STSP01 = self.manager.StsP01_text.upper()
        STSP02 = self.manager.StsP02_text.upper()
        STSP03 = self.manager.StsP03_text.upper()
        STSP04 = self.manager.StsP04_text.upper()
        STSP05 = self.manager.StsP05_text.upper()
        STSP06 = self.manager.StsP06_text.upper()
        STSP07 = self.manager.StsP07_text.upper()
        STSP08 = self.manager.StsP08_text.upper()
        STSP09 = self.manager.StsP09_text.upper()
        STSP10 = self.manager.StsP10_text.upper()
        STSP11 = self.manager.StsP11_text.upper()
        STSP12 = self.manager.StsP12_text.upper()
        STSP13 = self.manager.StsP13_text.upper()
        STSP14 = self.manager.StsP14_text.upper()
        STSP15 = self.manager.StsP15_text.upper()
        STSP16 = self.manager.StsP16_text.upper()

        cidade_text = self.manager.cidade_text.upper()
        pop_text = self.manager.pop_text.upper()
        slot_txt = self.manager.slot_text.upper()
        pon_text = self.manager.pon_text.upper()
        OBS_CTO = self.manager.obs_text

        if not hasattr(self.manager, 'OBS_CTO'):
            self.manager.OBS_CTO = " "

        cto_text = self.manager.cto_text.upper()
        cto_text = re.sub(r"\s+", "-", cto_text)

        A01 = f"{cidade_text}\n{pop_text}-SLOT:{slot_txt}-PON:{pon_text}-{cto_text} | 04-JOSÉ\n"
        A02 = f"P01 - {STSP01} - {time01} - {SNP01}"
        A03 = f"P02 - {STSP02} - {time02} - {SNP02}"
        A04 = f"P03 - {STSP03} - {time03} - {SNP03}"
        A05 = f"P04 - {STSP04} - {time04} - {SNP04}"
        A06 = f"P05 - {STSP05} - {time05} - {SNP05}"
        A07 = f"P06 - {STSP06} - {time06} - {SNP06}"
        A08 = f"P07 - {STSP07} - {time07} - {SNP07}"
        A09 = f"P08 - {STSP08} - {time08} - {SNP08}"
        A10 = f"P09 - {STSP09} - {time09} - {SNP09}"
        A11 = f"P10 - {STSP10} - {time10} - {SNP10}"
        A12 = f"P11 - {STSP11} - {time11} - {SNP11}"
        A13 = f"P12 - {STSP12} - {time12} - {SNP12}"
        A14 = f"P13 - {STSP13} - {time13} - {SNP13}"
        A15 = f"P14 - {STSP14} - {time14} - {SNP14}"
        A16 = f"P15 - {STSP15} - {time15} - {SNP15}"
        A17 = f"P16 - {STSP16} - {time16} - {SNP16}"
        A18 = f"\n{OBS_CTO}"

        all_portas = '\n'.join(
            [A01, A02, A03, A04, A05, A06, A07, A08, A09, A10, A11, A12, A13, A14, A15, A16, A17, A18])
        Clipboard.copy(all_portas)

        self.manager.current = 'Selecionar_spliter'
        self.manager.transition.direction = 'left'
        self.save_to_file2()

    def caixa_dialogo(self):
        import re
        from kivymd.uix.dialog import MDDialog
        from kivymd.uix.button import MDFlatButton

        if not hasattr(self.manager, 'SNP01_text'):
            self.manager.SNP01_text = "-"
        if not hasattr(self.manager, 'SNP02_text'):
            self.manager.SNP02_text = "-"
        if not hasattr(self.manager, 'SNP03_text'):
            self.manager.SNP03_text = "-"
        if not hasattr(self.manager, 'SNP04_text'):
            self.manager.SNP04_text = "-"
        if not hasattr(self.manager, 'SNP05_text'):
            self.manager.SNP05_text = "-"
        if not hasattr(self.manager, 'SNP06_text'):
            self.manager.SNP06_text = "-"
        if not hasattr(self.manager, 'SNP07_text'):
            self.manager.SNP07_text = "-"
        if not hasattr(self.manager, 'SNP08_text'):
            self.manager.SNP08_text = "-"
        if not hasattr(self.manager, 'SNP09_text'):
            self.manager.SNP09_text = "-"
        if not hasattr(self.manager, 'SNP10_text'):
            self.manager.SNP10_text = "-"
        if not hasattr(self.manager, 'SNP11_text'):
            self.manager.SNP11_text = "-"
        if not hasattr(self.manager, 'SNP12_text'):
            self.manager.SNP12_text = "-"
        if not hasattr(self.manager, 'SNP13_text'):
            self.manager.SNP13_text = "-"
        if not hasattr(self.manager, 'SNP14_text'):
            self.manager.SNP14_text = "-"
        if not hasattr(self.manager, 'SNP15_text'):
            self.manager.SNP15_text = "-"
        if not hasattr(self.manager, 'SNP16_text'):
            self.manager.SNP16_text = "-"

        if not hasattr(self.manager, 'StsP01_text'):
            self.manager.StsP01_text = "*"
        if not hasattr(self.manager, 'StsP02_text'):
            self.manager.StsP02_text = "*"
        if not hasattr(self.manager, 'StsP03_text'):
            self.manager.StsP03_text = "*"
        if not hasattr(self.manager, 'StsP04_text'):
            self.manager.StsP04_text = "*"
        if not hasattr(self.manager, 'StsP05_text'):
            self.manager.StsP05_text = "*"
        if not hasattr(self.manager, 'StsP06_text'):
            self.manager.StsP06_text = "*"
        if not hasattr(self.manager, 'StsP07_text'):
            self.manager.StsP07_text = "*"
        if not hasattr(self.manager, 'StsP08_text'):
            self.manager.StsP08_text = "*"
        if not hasattr(self.manager, 'StsP09_text'):
            self.manager.StsP09_text = "*"
        if not hasattr(self.manager, 'StsP10_text'):
            self.manager.StsP10_text = "*"
        if not hasattr(self.manager, 'StsP11_text'):
            self.manager.StsP11_text = "*"
        if not hasattr(self.manager, 'StsP12_text'):
            self.manager.StsP12_text = "*"
        if not hasattr(self.manager, 'StsP13_text'):
            self.manager.StsP13_text = "*"
        if not hasattr(self.manager, 'StsP14_text'):
            self.manager.StsP14_text = "*"
        if not hasattr(self.manager, 'StsP15_text'):
            self.manager.StsP15_text = "*"
        if not hasattr(self.manager, 'StsP16_text'):
            self.manager.StsP16_text = "*"

        time01 = self.time_text01
        time02 = self.time_text02
        time03 = self.time_text03
        time04 = self.time_text04
        time05 = self.time_text05
        time06 = self.time_text06
        time07 = self.time_text07
        time08 = self.time_text08
        time09 = self.time_text09
        time10 = self.time_text10
        time11 = self.time_text11
        time12 = self.time_text12
        time13 = self.time_text13
        time14 = self.time_text14
        time15 = self.time_text15
        time16 = self.time_text16

        SNP01 = self.manager.SNP01_text.upper()
        SNP02 = self.manager.SNP02_text.upper()
        SNP03 = self.manager.SNP03_text.upper()
        SNP04 = self.manager.SNP04_text.upper()
        SNP05 = self.manager.SNP05_text.upper()
        SNP06 = self.manager.SNP06_text.upper()
        SNP07 = self.manager.SNP07_text.upper()
        SNP08 = self.manager.SNP08_text.upper()
        SNP09 = self.manager.SNP09_text.upper()
        SNP10 = self.manager.SNP10_text.upper()
        SNP11 = self.manager.SNP11_text.upper()
        SNP12 = self.manager.SNP12_text.upper()
        SNP13 = self.manager.SNP13_text.upper()
        SNP14 = self.manager.SNP14_text.upper()
        SNP15 = self.manager.SNP15_text.upper()
        SNP16 = self.manager.SNP16_text.upper()

        STSP01 = self.manager.StsP01_text.upper()
        STSP02 = self.manager.StsP02_text.upper()
        STSP03 = self.manager.StsP03_text.upper()
        STSP04 = self.manager.StsP04_text.upper()
        STSP05 = self.manager.StsP05_text.upper()
        STSP06 = self.manager.StsP06_text.upper()
        STSP07 = self.manager.StsP07_text.upper()
        STSP08 = self.manager.StsP08_text.upper()
        STSP09 = self.manager.StsP09_text.upper()
        STSP10 = self.manager.StsP10_text.upper()
        STSP11 = self.manager.StsP11_text.upper()
        STSP12 = self.manager.StsP12_text.upper()
        STSP13 = self.manager.StsP13_text.upper()
        STSP14 = self.manager.StsP14_text.upper()
        STSP15 = self.manager.StsP15_text.upper()
        STSP16 = self.manager.StsP16_text.upper()

        porta01 = f"P01-{STSP01}-{time01}-{SNP01}"
        porta02 = f"P02-{STSP02}-{time02}-{SNP02}"
        porta03 = f"P03-{STSP03}-{time03}-{SNP03}"
        porta04 = f"P04-{STSP04}-{time04}-{SNP04}"
        porta05 = f"P05-{STSP05}-{time05}-{SNP05}"
        porta06 = f"P06-{STSP06}-{time06}-{SNP06}"
        porta07 = f"P07-{STSP07}-{time07}-{SNP07}"
        porta08 = f"P08-{STSP08}-{time08}-{SNP08}"
        porta09 = f"P09-{STSP09}-{time09}-{SNP09}"
        porta10 = f"P10-{STSP10}-{time10}-{SNP10}"
        porta11 = f"P11-{STSP11}-{time11}-{SNP11}"
        porta12 = f"P12-{STSP12}-{time12}-{SNP12}"
        porta13 = f"P13-{STSP13}-{time13}-{SNP13}"
        porta14 = f"P14-{STSP14}-{time14}-{SNP14}"
        porta15 = f"P15-{STSP15}-{time15}-{SNP15}"
        porta16 = f"P16-{STSP16}-{time16}-{SNP16}"

        cto_text = self.manager.cto_text.upper()
        cto_text = re.sub(r"\s+", "-", cto_text)
        OBS_CTO = self.manager.obs_text
        if not hasattr(self.manager, 'OBS_CTO'):
            self.manager.OBS_CTO = " "

        cidade_text = self.manager.cidade_text.upper()
        pop_text = self.manager.pop_text.upper()
        slot_txt = self.manager.slot_text
        pon_text = self.manager.pon_text

        dialogo = MDDialog(title=f"{cidade_text}\n\n{pop_text}-SLOT:{slot_txt}-PON:{pon_text}-{cto_text}",
                           text=f"\n{porta01}\n{porta02}\n{porta03}\n{porta04}\n{porta05}\n{porta06}\n{porta07}\n{porta08}\n{porta09}\n{porta10}\n{porta11}\n{porta12}\n{porta13}\n{porta14}\n{porta15}\n{porta16}\n\n{OBS_CTO}",
                           buttons=[
                               MDFlatButton(
                                   text="Fechar",
                                   theme_text_color="Custom",
                                   text_color=self.theme_cls.primary_color,
                                   on_release=lambda *args: self.caixa_de_dialogo.dismiss()
                               ),
                               MDFlatButton(
                                   text="Copiar",
                                   theme_text_color="Custom",
                                   text_color=self.theme_cls.primary_color,
                                   on_release=self.Copiar_Cto
                               ),
                               MDFlatButton(
                                   text="Salvar",
                                   theme_text_color="Custom",
                                   text_color=self.theme_cls.primary_color,
                                   on_release=self.salvar_progresso
                               ),

                           ],
                           )

        self.CX_dialogo = dialogo
        self.CX_dialogo.open()

    def tela_anterior(self):
        self.manager.current = 'Selecionar_spliter'
        self.manager.transition.direction = 'left'

    def tela_home(self):
        self.manager.current = 'HomeScreen'
        self.manager.transition.direction = 'left'

    def save_to_file2(self):
        import re
        import time
        import pandas as pd

        from android.permissions import request_permissions, Permission  # type: ignore
        from openpyxl import load_workbook

        if not hasattr(self.manager, 'StsP01_text'):
            self.manager.StsP01_text = "*"
        if not hasattr(self.manager, 'StsP02_text'):
            self.manager.StsP02_text = "*"
        if not hasattr(self.manager, 'StsP03_text'):
            self.manager.StsP03_text = "*"
        if not hasattr(self.manager, 'StsP04_text'):
            self.manager.StsP04_text = "*"
        if not hasattr(self.manager, 'StsP05_text'):
            self.manager.StsP05_text = "*"
        if not hasattr(self.manager, 'StsP06_text'):
            self.manager.StsP06_text = "*"
        if not hasattr(self.manager, 'StsP07_text'):
            self.manager.StsP07_text = "*"
        if not hasattr(self.manager, 'StsP08_text'):
            self.manager.StsP08_text = "*"
        if not hasattr(self.manager, 'StsP09_text'):
            self.manager.StsP09_text = "*"
        if not hasattr(self.manager, 'StsP10_text'):
            self.manager.StsP10_text = "*"
        if not hasattr(self.manager, 'StsP11_text'):
            self.manager.StsP11_text = "*"
        if not hasattr(self.manager, 'StsP12_text'):
            self.manager.StsP12_text = "*"
        if not hasattr(self.manager, 'StsP13_text'):
            self.manager.StsP13_text = "*"
        if not hasattr(self.manager, 'StsP14_text'):
            self.manager.StsP14_text = "*"
        if not hasattr(self.manager, 'StsP15_text'):
            self.manager.StsP15_text = "*"
        if not hasattr(self.manager, 'StsP16_text'):
            self.manager.StsP16_text = "*"

        if not hasattr(self.manager, 'SNP01_text'):
            self.manager.SNP01_text = "-"
        if not hasattr(self.manager, 'SNP02_text'):
            self.manager.SNP02_text = "-"
        if not hasattr(self.manager, 'SNP03_text'):
            self.manager.SNP03_text = "-"
        if not hasattr(self.manager, 'SNP04_text'):
            self.manager.SNP04_text = "-"
        if not hasattr(self.manager, 'SNP05_text'):
            self.manager.SNP05_text = "-"
        if not hasattr(self.manager, 'SNP06_text'):
            self.manager.SNP06_text = "-"
        if not hasattr(self.manager, 'SNP07_text'):
            self.manager.SNP07_text = "-"
        if not hasattr(self.manager, 'SNP08_text'):
            self.manager.SNP08_text = "-"
        if not hasattr(self.manager, 'SNP09_text'):
            self.manager.SNP09_text = "-"
        if not hasattr(self.manager, 'SNP10_text'):
            self.manager.SNP10_text = "-"
        if not hasattr(self.manager, 'SNP11_text'):
            self.manager.SNP11_text = "-"
        if not hasattr(self.manager, 'SNP12_text'):
            self.manager.SNP12_text = "-"
        if not hasattr(self.manager, 'SNP13_text'):
            self.manager.SNP13_text = "-"
        if not hasattr(self.manager, 'SNP14_text'):
            self.manager.SNP14_text = "-"
        if not hasattr(self.manager, 'SNP15_text'):
            self.manager.SNP15_text = "-"
        if not hasattr(self.manager, 'SNP16_text'):
            self.manager.SNP16_text = "-"

        if not hasattr(self.manager, 'PPP01_text'):
            self.manager.PPP01_text = "-"
        if not hasattr(self.manager, 'PPP02_text'):
            self.manager.PPP02_text = "-"
        if not hasattr(self.manager, 'PPP03_text'):
            self.manager.PPP03_text = "-"
        if not hasattr(self.manager, 'PPP04_text'):
            self.manager.PPP04_text = "-"
        if not hasattr(self.manager, 'PPP05_text'):
            self.manager.PPP05_text = "-"
        if not hasattr(self.manager, 'PPP06_text'):
            self.manager.PPP06_text = "-"
        if not hasattr(self.manager, 'PPP07_text'):
            self.manager.PPP07_text = "-"
        if not hasattr(self.manager, 'PPP08_text'):
            self.manager.PPP08_text = "-"
        if not hasattr(self.manager, 'PPP09_text'):
            self.manager.PPP09_text = "-"
        if not hasattr(self.manager, 'PPP10_text'):
            self.manager.PPP10_text = "-"
        if not hasattr(self.manager, 'PPP11_text'):
            self.manager.PPP11_text = "-"
        if not hasattr(self.manager, 'PPP12_text'):
            self.manager.PPP12_text = "-"
        if not hasattr(self.manager, 'PPP13_text'):
            self.manager.PPP13_text = "-"
        if not hasattr(self.manager, 'PPP14_text'):
            self.manager.PPP14_text = "-"
        if not hasattr(self.manager, 'PPP15_text'):
            self.manager.PPP15_text = "-"
        if not hasattr(self.manager, 'PPP16_text'):
            self.manager.PPP16_text = "-"

        if not hasattr(self.manager, 'Pwd01_text'):
            self.manager.Pwd01_text = "-"
        if not hasattr(self.manager, 'Pwd02_text'):
            self.manager.Pwd02_text = "-"
        if not hasattr(self.manager, 'Pwd03_text'):
            self.manager.Pwd03_text = "-"
        if not hasattr(self.manager, 'Pwd04_text'):
            self.manager.Pwd04_text = "-"
        if not hasattr(self.manager, 'Pwd05_text'):
            self.manager.Pwd05_text = "-"
        if not hasattr(self.manager, 'Pwd06_text'):
            self.manager.Pwd06_text = "-"
        if not hasattr(self.manager, 'Pwd07_text'):
            self.manager.Pwd07_text = "-"
        if not hasattr(self.manager, 'Pwd08_text'):
            self.manager.Pwd08_text = "-"
        if not hasattr(self.manager, 'Pwd09_text'):
            self.manager.Pwd09_text = "-"
        if not hasattr(self.manager, 'Pwd10_text'):
            self.manager.Pwd10_text = "-"
        if not hasattr(self.manager, 'Pwd11_text'):
            self.manager.Pwd11_text = "-"
        if not hasattr(self.manager, 'Pwd12_text'):
            self.manager.Pwd12_text = "-"
        if not hasattr(self.manager, 'Pwd13_text'):
            self.manager.Pwd13_text = "-"
        if not hasattr(self.manager, 'Pwd14_text'):
            self.manager.Pwd14_text = "-"
        if not hasattr(self.manager, 'Pwd15_text'):
            self.manager.Pwd15_text = "-"
        if not hasattr(self.manager, 'Pwd16_text'):
            self.manager.Pwd16_text = "-"

        if not hasattr(self.manager, 'Loid01_text'):
            self.manager.Loid01_text = "-"
        if not hasattr(self.manager, 'Loid02_text'):
            self.manager.Loid02_text = "-"
        if not hasattr(self.manager, 'Loid03_text'):
            self.manager.Loid03_text = "-"
        if not hasattr(self.manager, 'Loid04_text'):
            self.manager.Loid04_text = "-"
        if not hasattr(self.manager, 'Loid05_text'):
            self.manager.Loid05_text = "-"
        if not hasattr(self.manager, 'Loid06_text'):
            self.manager.Loid06_text = "-"
        if not hasattr(self.manager, 'Loid07_text'):
            self.manager.Loid07_text = "-"
        if not hasattr(self.manager, 'Loid08_text'):
            self.manager.Loid08_text = "-"
        if not hasattr(self.manager, 'Loid09_text'):
            self.manager.Loid09_text = "-"
        if not hasattr(self.manager, 'Loid10_text'):
            self.manager.Loid10_text = "-"
        if not hasattr(self.manager, 'Loid11_text'):
            self.manager.Loid11_text = "-"
        if not hasattr(self.manager, 'Loid12_text'):
            self.manager.Loid12_text = "-"
        if not hasattr(self.manager, 'Loid13_text'):
            self.manager.Loid13_text = "-"
        if not hasattr(self.manager, 'Loid14_text'):
            self.manager.Loid14_text = "-"
        if not hasattr(self.manager, 'Loid15_text'):
            self.manager.Loid15_text = "-"
        if not hasattr(self.manager, 'Loid16_text'):
            self.manager.Loid16_text = "-"

        if not hasattr(self.manager, 'Obs01_text'):
            self.manager.Obs01_text = "-"
        if not hasattr(self.manager, 'Obs02_text'):
            self.manager.Obs02_text = "-"
        if not hasattr(self.manager, 'Obs03_text'):
            self.manager.Obs03_text = "-"
        if not hasattr(self.manager, 'Obs04_text'):
            self.manager.Obs04_text = "-"
        if not hasattr(self.manager, 'Obs05_text'):
            self.manager.Obs05_text = "-"
        if not hasattr(self.manager, 'Obs06_text'):
            self.manager.Obs06_text = "-"
        if not hasattr(self.manager, 'Obs07_text'):
            self.manager.Obs07_text = "-"
        if not hasattr(self.manager, 'Obs08_text'):
            self.manager.Obs08_text = "-"
        if not hasattr(self.manager, 'Obs09_text'):
            self.manager.Obs09_text = "-"
        if not hasattr(self.manager, 'Obs10_text'):
            self.manager.Obs10_text = "-"
        if not hasattr(self.manager, 'Obs11_text'):
            self.manager.Obs11_text = "-"
        if not hasattr(self.manager, 'Obs12_text'):
            self.manager.Obs12_text = "-"
        if not hasattr(self.manager, 'Obs13_text'):
            self.manager.Obs13_text = "-"
        if not hasattr(self.manager, 'Obs14_text'):
            self.manager.Obs14_text = "-"
        if not hasattr(self.manager, 'Obs15_text'):
            self.manager.Obs15_text = "-"
        if not hasattr(self.manager, 'Obs16_text'):
            self.manager.Obs16_text = "-"

        StsP01 = self.manager.StsP01_text.upper()
        StsP02 = self.manager.StsP02_text.upper()
        StsP03 = self.manager.StsP03_text.upper()
        StsP04 = self.manager.StsP04_text.upper()
        StsP05 = self.manager.StsP05_text.upper()
        StsP06 = self.manager.StsP06_text.upper()
        StsP07 = self.manager.StsP07_text.upper()
        StsP08 = self.manager.StsP08_text.upper()
        StsP09 = self.manager.StsP09_text.upper()
        StsP10 = self.manager.StsP10_text.upper()
        StsP11 = self.manager.StsP11_text.upper()
        StsP12 = self.manager.StsP12_text.upper()
        StsP13 = self.manager.StsP13_text.upper()
        StsP14 = self.manager.StsP14_text.upper()
        StsP15 = self.manager.StsP15_text.upper()
        StsP16 = self.manager.StsP16_text.upper()

        SNP01 = self.manager.SNP01_text.upper()
        SNP02 = self.manager.SNP02_text.upper()
        SNP03 = self.manager.SNP03_text.upper()
        SNP04 = self.manager.SNP04_text.upper()
        SNP05 = self.manager.SNP05_text.upper()
        SNP06 = self.manager.SNP06_text.upper()
        SNP07 = self.manager.SNP07_text.upper()
        SNP08 = self.manager.SNP08_text.upper()
        SNP09 = self.manager.SNP09_text.upper()
        SNP10 = self.manager.SNP10_text.upper()
        SNP11 = self.manager.SNP11_text.upper()
        SNP12 = self.manager.SNP12_text.upper()
        SNP13 = self.manager.SNP13_text.upper()
        SNP14 = self.manager.SNP14_text.upper()
        SNP15 = self.manager.SNP15_text.upper()
        SNP16 = self.manager.SNP16_text.upper()

        PPP01 = self.manager.PPP01_text.upper()
        PPP02 = self.manager.PPP02_text.upper()
        PPP03 = self.manager.PPP03_text.upper()
        PPP04 = self.manager.PPP04_text.upper()
        PPP05 = self.manager.PPP05_text.upper()
        PPP06 = self.manager.PPP06_text.upper()
        PPP07 = self.manager.PPP07_text.upper()
        PPP08 = self.manager.PPP08_text.upper()
        PPP09 = self.manager.PPP09_text.upper()
        PPP10 = self.manager.PPP10_text.upper()
        PPP11 = self.manager.PPP11_text.upper()
        PPP12 = self.manager.PPP12_text.upper()
        PPP13 = self.manager.PPP13_text.upper()
        PPP14 = self.manager.PPP14_text.upper()
        PPP15 = self.manager.PPP15_text.upper()
        PPP16 = self.manager.PPP16_text.upper()

        Pwd01 = self.manager.Pwd01_text.upper()
        Pwd02 = self.manager.Pwd02_text.upper()
        Pwd03 = self.manager.Pwd03_text.upper()
        Pwd04 = self.manager.Pwd04_text.upper()
        Pwd05 = self.manager.Pwd05_text.upper()
        Pwd06 = self.manager.Pwd06_text.upper()
        Pwd07 = self.manager.Pwd07_text.upper()
        Pwd08 = self.manager.Pwd08_text.upper()
        Pwd09 = self.manager.Pwd09_text.upper()
        Pwd10 = self.manager.Pwd10_text.upper()
        Pwd11 = self.manager.Pwd11_text.upper()
        Pwd12 = self.manager.Pwd12_text.upper()
        Pwd13 = self.manager.Pwd13_text.upper()
        Pwd14 = self.manager.Pwd14_text.upper()
        Pwd15 = self.manager.Pwd15_text.upper()
        Pwd16 = self.manager.Pwd16_text.upper()

        Loid01 = self.manager.Loid01_text.upper()
        Loid02 = self.manager.Loid02_text.upper()
        Loid03 = self.manager.Loid03_text.upper()
        Loid04 = self.manager.Loid04_text.upper()
        Loid05 = self.manager.Loid05_text.upper()
        Loid06 = self.manager.Loid06_text.upper()
        Loid07 = self.manager.Loid07_text.upper()
        Loid08 = self.manager.Loid08_text.upper()
        Loid09 = self.manager.Loid09_text.upper()
        Loid10 = self.manager.Loid10_text.upper()
        Loid11 = self.manager.Loid11_text.upper()
        Loid12 = self.manager.Loid12_text.upper()
        Loid13 = self.manager.Loid13_text.upper()
        Loid14 = self.manager.Loid14_text.upper()
        Loid15 = self.manager.Loid15_text.upper()
        Loid16 = self.manager.Loid16_text.upper()

        Obs01 = self.manager.Obs01_text.upper()
        Obs02 = self.manager.Obs02_text.upper()
        Obs03 = self.manager.Obs03_text.upper()
        Obs04 = self.manager.Obs04_text.upper()
        Obs05 = self.manager.Obs05_text.upper()
        Obs06 = self.manager.Obs06_text.upper()
        Obs07 = self.manager.Obs07_text.upper()
        Obs08 = self.manager.Obs08_text.upper()
        Obs09 = self.manager.Obs09_text.upper()
        Obs10 = self.manager.Obs10_text.upper()
        Obs11 = self.manager.Obs11_text.upper()
        Obs12 = self.manager.Obs12_text.upper()
        Obs13 = self.manager.Obs13_text.upper()
        Obs14 = self.manager.Obs14_text.upper()
        Obs15 = self.manager.Obs15_text.upper()
        Obs16 = self.manager.Obs16_text.upper()

        ST01 = self.start_time01
        ST02 = self.start_time02
        ST03 = self.start_time03
        ST04 = self.start_time04
        ST05 = self.start_time05
        ST06 = self.start_time06
        ST07 = self.start_time07
        ST08 = self.start_time08
        ST09 = self.start_time09
        ST10 = self.start_time10
        ST11 = self.start_time11
        ST12 = self.start_time12
        ST13 = self.start_time13
        ST14 = self.start_time14
        ST15 = self.start_time15
        ST16 = self.start_time16

        ET01 = self.end_time01
        ET02 = self.end_time02
        ET03 = self.end_time03
        ET04 = self.end_time04
        ET05 = self.end_time05
        ET06 = self.end_time06
        ET07 = self.end_time07
        ET08 = self.end_time08
        ET09 = self.end_time09
        ET10 = self.end_time10
        ET11 = self.end_time11
        ET12 = self.end_time12
        ET13 = self.end_time13
        ET14 = self.end_time14
        ET15 = self.end_time15
        ET16 = self.end_time16

        slot_txt = self.manager.slot_text.upper()
        pon_text = self.manager.pon_text.upper()
        cto_text = self.manager.cto_text.upper()

        slotpon = f'SLOT {slot_txt} - PON {pon_text}'
        slotpon2 = f'{slot_txt}/{pon_text}'
        objetos = [StsP01, StsP02, StsP03, StsP04, StsP05, StsP06, StsP07, StsP08, StsP09, StsP10, StsP11, StsP12, StsP13, StsP14, StsP15, StsP16]
        cto_text = re.sub(r"\s+", "-", cto_text)
        OBS_CTO = self.manager.obs_text
        if not hasattr(self.manager, 'OBS_CTO'):
            self.manager.OBS_CTO = " "

        contador_hashtag = 0
        contador_ok = 0

        for obj in objetos:
            contador_hashtag += obj.count("#")
            contador_ok += obj.count("OK")

        PVG = contador_hashtag + contador_ok

        cidade_text = self.manager.cidade_text.upper()
        pop_text = self.manager.pop_text.upper()
        slot_txt = self.manager.slot_text.upper()
        pon_text = self.manager.pon_text.upper()
        data_hoje = time.strftime("%d/%m/%Y")

        id_tec = "04 - JOSÉ"

        import os
        from plyer import storagepath
        from android.permissions import request_permissions, Permission  # type: ignore
        request_permissions([Permission.READ_EXTERNAL_STORAGE, Permission.WRITE_EXTERNAL_STORAGE, ])

        documents_path = storagepath.get_documents_dir()
        folder_path = os.path.join(documents_path,  'FinderPlanilha', f'{cidade_text}', f'{pop_text}', f'{slotpon}',
                                   'Planilha')
        os.makedirs(folder_path, exist_ok=True)
        file_path = os.path.join(folder_path, f'{cidade_text}-{pop_text}-{slotpon}.xlsx')

        CTO_PLANILHA = {
            'DATA|CTO': [data_hoje, cto_text, cto_text, cto_text, cto_text, cto_text, cto_text, cto_text,
                    cto_text, cto_text, cto_text, cto_text, cto_text, cto_text, cto_text, cto_text, cto_text, " "],
            'EQUIPE|PORTA': [id_tec, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, " "],
            'CIDADE|STATUS': [cidade_text, StsP01, StsP02, StsP03, StsP04, StsP05, StsP06, StsP07, StsP08,
                       StsP09, StsP10, StsP11, StsP12, StsP13, StsP14, StsP15, StsP16, " "],
            'POP|DESC.': [pop_text, ST01, ST02, ST03, ST04, ST05, ST06, ST07, ST08, ST09, ST10, ST11, ST12, ST13,
                           ST14, ST15, ST16, " "],
            'SLOT|CONEX.': [slot_txt, ET01, ET02, ET03, ET04, ET05, ET06, ET07, ET08, ET09, ET10, ET11, ET12, ET13,
                        ET14, ET15, ET16, " "],
            'PON|FSAN/SN': [pon_text, SNP01, SNP02, SNP03, SNP04, SNP05, SNP06, SNP07, SNP08, SNP09, SNP10,
                        SNP11, SNP12, SNP13, SNP14, SNP15, SNP16, " "],
            'PPPoE': ["-", PPP01, PPP02, PPP03, PPP04, PPP05, PPP06, PPP07, PPP08, PPP09, PPP10, PPP11, PPP12,
                      PPP13, PPP14, PPP15, PPP16, " "],
            'PWD': ["-", Pwd01, Pwd02, Pwd03, Pwd04, Pwd05, Pwd06, Pwd07, Pwd08, Pwd09, Pwd10, Pwd11, Pwd12, Pwd13,
                    Pwd14, Pwd15, Pwd16, " "],
            'LOID': ["-", Loid01, Loid02, Loid03, Loid04, Loid05, Loid06, Loid07, Loid08, Loid09, Loid10, Loid11,
                     Loid12, Loid13, Loid14, Loid15, Loid16, " "],
            'TOTAL CLIENTES|OBS.': [PVG, Obs01, Obs02, Obs03, Obs04, Obs05, Obs06, Obs07, Obs08, Obs09, Obs10,
                    Obs11, Obs12, Obs13, Obs14, Obs15, Obs16 + OBS_CTO, " "],
        }

        CTO_RESUMO = {
            'SPLITAGEM': ["1/8"],
            'CTOs': [cto_text],
            'SLOT/PON': [slotpon2],
            'EQUIPE': [id_tec],
            'DROPs REDEX': [contador_ok],
            'DROPs ATIVOS': [" "],
            'DROPs N/I REDEX': [contador_hashtag],
            'TOTAL DROPs P/ CTO': [PVG],
            'TOTAL DROPs N/I': [' '],
            'DATA': [data_hoje],
        }

        df1 = pd.DataFrame(CTO_PLANILHA)
        df2 = pd.DataFrame(CTO_RESUMO)

        if os.path.exists(file_path):
            book = load_workbook(file_path)
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                writer.book = book
                writer.sheets = {ws.title: ws for ws in book.worksheets}
                if 'CTO_PLANILHA' in writer.sheets:
                    startrow1 = writer.sheets['CTO_PLANILHA'].max_row
                else:
                    startrow1 = 0

                if 'CTO_RESUMO' in writer.sheets:
                    startrow2 = writer.sheets['CTO_RESUMO'].max_row
                else:
                    startrow2 = 0

                df1.to_excel(writer, sheet_name='CTO_PLANILHA', startrow=startrow1, index=False, header=False)
                df2.to_excel(writer, sheet_name='CTO_RESUMO', startrow=startrow2, index=False, header=False)
        else:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df1.to_excel(writer, sheet_name='CTO_PLANILHA', index=False)
                df2.to_excel(writer, sheet_name='CTO_RESUMO', index=False)

        from openpyxl import load_workbook
        from openpyxl.styles import Border, Side

        # Carregue o arquivo XLSX existente
        book = load_workbook(file_path)

        # Selecione a planilha desejada (por exemplo, 'CTO_PLANILHA')
        sheet1 = book['CTO_PLANILHA']
        sheet1.column_dimensions['A'].width = 11
        sheet1.column_dimensions['B'].width = 15
        sheet1.column_dimensions['C'].width = 15
        sheet1.column_dimensions['D'].width = 12
        sheet1.column_dimensions['E'].width = 12
        sheet1.column_dimensions['F'].width = 20
        sheet1.column_dimensions['G'].width = 10
        sheet1.column_dimensions['H'].width = 10
        sheet1.column_dimensions['I'].width = 10
        sheet1.column_dimensions['J'].width = 20

        sheet2 = book['CTO_RESUMO']
        sheet2.column_dimensions['A'].width = 12
        sheet2.column_dimensions['B'].width = 17
        sheet2.column_dimensions['C'].width = 16
        sheet2.column_dimensions['D'].width = 13
        sheet2.column_dimensions['E'].width = 15
        sheet2.column_dimensions['F'].width = 18
        sheet2.column_dimensions['G'].width = 10
        sheet2.column_dimensions['H'].width = 10
        sheet2.column_dimensions['I'].width = 10
        sheet2.column_dimensions['J'].width = 25

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        sheet1['A1'].fill = PatternFill(fill_type='solid', start_color='5cb800')
        sheet1['B1'].fill = PatternFill(fill_type='solid', start_color='5cb800')
        sheet1['C1'].fill = PatternFill(fill_type='solid', start_color='5cb800')
        sheet1['D1'].fill = PatternFill(fill_type='solid', start_color='5cb800')
        sheet1['E1'].fill = PatternFill(fill_type='solid', start_color='5cb800')
        sheet1['F1'].fill = PatternFill(fill_type='solid', start_color='5cb800')
        sheet1['G1'].fill = PatternFill(fill_type='solid', start_color='5cb800')
        sheet1['H1'].fill = PatternFill(fill_type='solid', start_color='5cb800')
        sheet1['I1'].fill = PatternFill(fill_type='solid', start_color='5cb800')
        sheet1['J1'].fill = PatternFill(fill_type='solid', start_color='5cb800')

        # Percorra as células e aplique a borda
        for row in sheet1.iter_rows():
            for cell in row:
                cell.border = thin_border

        for row in sheet2.iter_rows():
            for cell in row:
                cell.border = thin_border

        # Salve o arquivo
        book.save(file_path)

        self.Limpar_widget_1x16()
        self.Limpar_Hora()

    def Limpar_widget_1x16(self):
        self.manager.get_screen('PB01').clear_text()
        self.manager.get_screen('PB02').clear_text()
        self.manager.get_screen('PB03').clear_text()
        self.manager.get_screen('PB04').clear_text()
        self.manager.get_screen('PB05').clear_text()
        self.manager.get_screen('PB06').clear_text()
        self.manager.get_screen('PB07').clear_text()
        self.manager.get_screen('PB08').clear_text()
        self.manager.get_screen('PB09').clear_text()
        self.manager.get_screen('PB10').clear_text()
        self.manager.get_screen('PB11').clear_text()
        self.manager.get_screen('PB12').clear_text()
        self.manager.get_screen('PB13').clear_text()
        self.manager.get_screen('PB14').clear_text()
        self.manager.get_screen('PB15').clear_text()
        self.manager.get_screen('PB16').clear_text()

    def criarpasta(self):
        import os

        from android.permissions import request_permissions, Permission  # type: ignore
        request_permissions([Permission.READ_EXTERNAL_STORAGE, Permission.WRITE_EXTERNAL_STORAGE, ])
        from plyer import storagepath
        documents_path = storagepath.get_documents_dir()
        folder_path = os.path.join(documents_path, 'FinderPlanilha', 'Rioclaro')
        os.makedirs(folder_path, exist_ok=True)


class PA01(MDScreen):

    def check_box_status(self):
        if self.ids.statusok.active:
            self.StsP01 = 'OK'
            print(f"STATUS:{self.StsP01}")
        elif self.ids.statusbko.active:
            self.StsP01 = '#'
            print(f"STATUS:{self.StsP01}")
        else:
            self.StsP01 = '*'
            print(f"STATUS:{self.StsP01}")

    def clear_text(self):
        self.ids.SNP01.text = ""
        self.ids.PPP01.text = ""
        self.ids.Pwd01.text = ""
        self.ids.Obs01.text = ""
        self.ids.Loid01.text = ""

    def Salvar_Info_PA01(self):
        self.manager.current = "tela1x8"
        self.manager.transition.direction = 'left'
        self.check_box_status()
        self.manager.StsP01_text = self.StsP01.upper()
        self.manager.Loid01_text = self.ids.Loid01.text.upper()
        self.manager.SNP01_text = self.ids.SNP01.text.upper()
        self.manager.PPP01_text = self.ids.PPP01.text.upper()
        self.manager.Pwd01_text = self.ids.Pwd01.text.upper()
        self.manager.Obs01_text = self.ids.Obs01.text.upper()


class PA02(MDScreen):
    def check_box_status(self):
        if self.ids.statusok.active:
            self.StsP02 = 'OK'
            print(f"STATUS:{self.StsP02}")
        elif self.ids.statusbko.active:
            self.StsP02 = '#'
            print(f"STATUS:{self.StsP02}")
        else:
            self.StsP02 = '*'
            print(f"STATUS:{self.StsP02}")

    def clear_text(self):
        self.ids.SNP02.text = ""
        self.ids.PPP02.text = ""
        self.ids.Pwd02.text = ""
        self.ids.Obs02.text = ""
        self.ids.Loid02.text = ""

    def Salvar_Info_PA02(self):
        self.manager.current = "tela1x8"
        self.manager.transition.direction = 'left'
        self.check_box_status()
        self.manager.StsP02_text = self.StsP02.upper()
        self.manager.Loid02_text = self.ids.Loid02.text.upper()
        self.manager.SNP02_text = self.ids.SNP02.text.upper()
        self.manager.PPP02_text = self.ids.PPP02.text.upper()
        self.manager.Pwd02_text = self.ids.Pwd02.text.upper()
        self.manager.Obs02_text = self.ids.Obs02.text.upper()


class PA03(MDScreen):
    def check_box_status(self):
        if self.ids.statusok.active:
            self.StsP03 = 'OK'
            print(f"STATUS:{self.StsP03}")
        elif self.ids.statusbko.active:
            self.StsP03 = '#'
            print(f"STATUS:{self.StsP03}")
        else:
            self.StsP03 = '*'
            print(f"STATUS:{self.StsP03}")

    def clear_text(self):
        self.ids.SNP03.text = ""
        self.ids.PPP03.text = ""
        self.ids.Pwd03.text = ""
        self.ids.Obs03.text = ""
        self.ids.Loid03.text = ""

    def Salvar_Info_PA03(self):
        self.manager.current = "tela1x8"
        self.manager.transition.direction = 'left'
        self.check_box_status()

        self.manager.StsP03_text = self.StsP03.upper()
        self.manager.Loid03_text = self.ids.Loid03.text.upper()
        self.manager.SNP03_text = self.ids.SNP03.text.upper()
        self.manager.PPP03_text = self.ids.PPP03.text.upper()
        self.manager.Pwd03_text = self.ids.Pwd03.text.upper()
        self.manager.Obs03_text = self.ids.Obs03.text.upper()


class PA04(MDScreen):
    def check_box_status(self):
        if self.ids.statusok.active:
            self.StsP04 = 'OK'
            print(f"STATUS:{self.StsP04}")
        elif self.ids.statusbko.active:
            self.StsP04 = '#'
            print(f"STATUS:{self.StsP04}")
        else:
            self.StsP04 = '*'
            print(f"STATUS:{self.StsP04}")

    def clear_text(self):
        self.ids.SNP04.text = ""
        self.ids.PPP04.text = ""
        self.ids.Pwd04.text = ""
        self.ids.Obs04.text = ""
        self.ids.Loid04.text = ""

    def Salvar_Info_PA04(self):
        self.manager.current = "tela1x8"
        self.manager.transition.direction = 'left'
        self.check_box_status()

        self.manager.StsP04_text = self.StsP04.upper()
        self.manager.Loid04_text = self.ids.Loid04.text.upper()
        self.manager.SNP04_text = self.ids.SNP04.text.upper()
        self.manager.PPP04_text = self.ids.PPP04.text.upper()
        self.manager.Pwd04_text = self.ids.Pwd04.text.upper()
        self.manager.Obs04_text = self.ids.Obs04.text.upper()


class PA05(MDScreen):
    def check_box_status(self):
        if self.ids.statusok.active:
            self.StsP05 = 'OK'
            print(f"STATUS:{self.StsP05}")
        elif self.ids.statusbko.active:
            self.StsP05 = '#'
            print(f"STATUS:{self.StsP05}")
        else:
            self.StsP05 = '*'
            print(f"STATUS:{self.StsP05}")

    def clear_text(self):
        self.ids.SNP05.text = ""
        self.ids.PPP05.text = ""
        self.ids.Pwd05.text = ""
        self.ids.Obs05.text = ""
        self.ids.Loid05.text = ""

    def Salvar_Info_PA05(self):
        self.manager.current = "tela1x8"
        self.manager.transition.direction = 'left'
        self.check_box_status()

        self.manager.StsP05_text = self.StsP05.upper()
        self.manager.Loid05_text = self.ids.Loid05.text.upper()
        self.manager.SNP05_text = self.ids.SNP05.text.upper()
        self.manager.PPP05_text = self.ids.PPP05.text.upper()
        self.manager.Pwd05_text = self.ids.Pwd05.text.upper()
        self.manager.Obs05_text = self.ids.Obs05.text.upper()


class PA06(MDScreen):
    def check_box_status(self):
        if self.ids.statusok.active:
            self.StsP06 = 'OK'
            print(f"STATUS:{self.StsP06}")
        elif self.ids.statusbko.active:
            self.StsP06 = '#'
            print(f"STATUS:{self.StsP06}")
        else:
            self.StsP06 = '*'
            print(f"STATUS:{self.StsP06}")

    def clear_text(self):
        self.ids.SNP06.text = ""
        self.ids.PPP06.text = ""
        self.ids.Pwd06.text = ""
        self.ids.Obs06.text = ""
        self.ids.Loid06.text = ""

    def Salvar_Info_PA06(self):
        self.manager.current = "tela1x8"
        self.manager.transition.direction = 'left'
        self.check_box_status()

        self.manager.StsP06_text = self.StsP06.upper()
        self.manager.Loid06_text = self.ids.Loid06.text.upper()
        self.manager.SNP06_text = self.ids.SNP06.text.upper()
        self.manager.PPP06_text = self.ids.PPP06.text.upper()
        self.manager.Pwd06_text = self.ids.Pwd06.text.upper()
        self.manager.Obs06_text = self.ids.Obs06.text.upper()


class PA07(MDScreen):
    def check_box_status(self):
        if self.ids.statusok.active:
            self.StsP07 = 'OK'
            print(f"STATUS:{self.StsP07}")
        elif self.ids.statusbko.active:
            self.StsP07 = '#'
            print(f"STATUS:{self.StsP07}")
        else:
            self.StsP07 = '*'
            print(f"STATUS:{self.StsP07}")

    def clear_text(self):
        self.ids.SNP07.text = ""
        self.ids.PPP07.text = ""
        self.ids.Pwd07.text = ""
        self.ids.Obs07.text = ""
        self.ids.Loid07.text = ""

    def Salvar_Info_PA07(self):
        self.manager.current = "tela1x8"
        self.manager.transition.direction = 'left'
        self.check_box_status()

        self.manager.StsP07_text = self.StsP07.upper()
        self.manager.Loid07_text = self.ids.Loid07.text.upper()
        self.manager.SNP07_text = self.ids.SNP07.text.upper()
        self.manager.PPP07_text = self.ids.PPP07.text.upper()
        self.manager.Pwd07_text = self.ids.Pwd07.text.upper()
        self.manager.Obs07_text = self.ids.Obs07.text.upper()


class PA08(MDScreen):
    def check_box_status(self):
        if self.ids.statusok.active:
            self.StsP08 = 'OK'
            print(f"STATUS:{self.StsP08}")
        elif self.ids.statusbko.active:
            self.StsP08 = '#'
            print(f"STATUS:{self.StsP08}")
        else:
            self.StsP08 = '*'
            print(f"STATUS:{self.StsP08}")

    def clear_text(self):
        self.ids.SNP08.text = ""
        self.ids.PPP08.text = ""
        self.ids.Pwd08.text = ""
        self.ids.Obs08.text = ""
        self.ids.Loid08.text = ""

    def Salvar_Info_PA08(self):
        self.manager.current = "tela1x8"
        self.manager.transition.direction = 'left'
        self.check_box_status()

        self.manager.StsP08_text = self.StsP08.upper()
        self.manager.Loid08_text = self.ids.Loid08.text.upper()
        self.manager.SNP08_text = self.ids.SNP08.text.upper()
        self.manager.PPP08_text = self.ids.PPP08.text.upper()
        self.manager.Pwd08_text = self.ids.Pwd08.text.upper()
        self.manager.Obs08_text = self.ids.Obs08.text.upper()


class PB01(MDScreen):
    def clear_text(self):
        self.ids.SNP01.text = ""
        self.ids.PPP01.text = ""
        self.ids.Pwd01.text = ""
        self.ids.Obs01.text = ""
        self.ids.Loid01.text = ""

    def check_box_status(self):
        if self.ids.statusok.active:
            self.StsP01 = 'OK'
            print(f"STATUS:{self.StsP01}")
        elif self.ids.statusbko.active:
            self.StsP01 = '#'
            print(f"STATUS:{self.StsP01}")
        else:
            self.StsP01 = '*'
            print(f"STATUS:{self.StsP01}")

    def Salvar_Info_PB01(self):
        self.manager.current = "tela1x16"
        self.manager.transition.direction = 'left'
        self.check_box_status()

        self.manager.StsP01_text = self.StsP01.upper()
        self.manager.Loid01_text = self.ids.Loid01.text.upper()
        self.manager.SNP01_text = self.ids.SNP01.text.upper()
        self.manager.PPP01_text = self.ids.PPP01.text.upper()
        self.manager.Pwd01_text = self.ids.Pwd01.text.upper()
        self.manager.Obs01_text = self.ids.Obs01.text.upper()


class PB02(MDScreen):
    def clear_text(self):
        self.ids.SNP02.text = ""
        self.ids.PPP02.text = ""
        self.ids.Pwd02.text = ""
        self.ids.Obs02.text = ""
        self.ids.Loid02.text = ""

    def check_box_status(self):
        if self.ids.statusok.active:
            self.StsP02 = 'OK'
            print(f"STATUS:{self.StsP02}")
        elif self.ids.statusbko.active:
            self.StsP02 = '#'
            print(f"STATUS:{self.StsP02}")
        else:
            self.StsP02 = '*'
            print(f"STATUS:{self.StsP02}")

    def Salvar_Info_PB02(self):
        self.manager.current = "tela1x16"
        self.manager.transition.direction = 'left'
        self.check_box_status()

        self.manager.StsP02_text = self.StsP02.upper()
        self.manager.Loid02_text = self.ids.Loid02.text.upper()
        self.manager.SNP02_text = self.ids.SNP02.text.upper()
        self.manager.PPP02_text = self.ids.PPP02.text.upper()
        self.manager.Pwd02_text = self.ids.Pwd02.text.upper()
        self.manager.Obs02_text = self.ids.Obs02.text.upper()


class PB03(MDScreen):
    def clear_text(self):
        self.ids.SNP03.text = ""
        self.ids.PPP03.text = ""
        self.ids.Pwd03.text = ""
        self.ids.Obs03.text = ""
        self.ids.Loid03.text = ""

    def check_box_status(self):
        if self.ids.statusok.active:
            self.StsP03 = 'OK'
            print(f"STATUS:{self.StsP03}")
        elif self.ids.statusbko.active:
            self.StsP03 = '#'
            print(f"STATUS:{self.StsP03}")
        else:
            self.StsP03 = '*'
            print(f"STATUS:{self.StsP03}")

    def Salvar_Info_PB03(self):
        self.manager.current = "tela1x16"
        self.manager.transition.direction = 'left'
        self.check_box_status()

        self.manager.StsP03_text = self.StsP03.upper()
        self.manager.Loid03_text = self.ids.Loid03.text.upper()
        self.manager.SNP03_text = self.ids.SNP03.text.upper()
        self.manager.PPP03_text = self.ids.PPP03.text.upper()
        self.manager.Pwd03_text = self.ids.Pwd03.text.upper()
        self.manager.Obs03_text = self.ids.Obs03.text.upper()


class PB04(MDScreen):
    def clear_text(self):
        self.ids.SNP04.text = ""
        self.ids.PPP04.text = ""
        self.ids.Pwd04.text = ""
        self.ids.Obs04.text = ""
        self.ids.Loid04.text = ""

    def check_box_status(self):
        if self.ids.statusok.active:
            self.StsP04 = 'OK'
            print(f"STATUS:{self.StsP04}")
        elif self.ids.statusbko.active:
            self.StsP04 = '#'
            print(f"STATUS:{self.StsP04}")
        else:
            self.StsP04 = '*'
            print(f"STATUS:{self.StsP04}")

    def Salvar_Info_PB04(self):
        self.manager.current = "tela1x16"
        self.manager.transition.direction = 'left'
        self.check_box_status()

        self.manager.StsP04_text = self.StsP04.upper()
        self.manager.Loid04_text = self.ids.Loid04.text.upper()
        self.manager.SNP04_text = self.ids.SNP04.text.upper()
        self.manager.PPP04_text = self.ids.PPP04.text.upper()
        self.manager.Pwd04_text = self.ids.Pwd04.text.upper()
        self.manager.Obs04_text = self.ids.Obs04.text.upper()


class PB05(MDScreen):
    def clear_text(self):
        self.ids.SNP05.text = ""
        self.ids.PPP05.text = ""
        self.ids.Pwd05.text = ""
        self.ids.Obs05.text = ""
        self.ids.Loid05.text = ""

    def check_box_status(self):
        if self.ids.statusok.active:
            self.StsP05 = 'OK'
            print(f"STATUS:{self.StsP05}")
        elif self.ids.statusbko.active:
            self.StsP05 = '#'
            print(f"STATUS:{self.StsP05}")
        else:
            self.StsP05 = '*'
            print(f"STATUS:{self.StsP05}")

    def Salvar_Info_PB05(self):
        self.manager.current = "tela1x16"
        self.manager.transition.direction = 'left'
        self.check_box_status()

        self.manager.StsP05_text = self.StsP05.upper()
        self.manager.Loid05_text = self.ids.Loid05.text.upper()
        self.manager.SNP05_text = self.ids.SNP05.text.upper()
        self.manager.PPP05_text = self.ids.PPP05.text.upper()
        self.manager.Pwd05_text = self.ids.Pwd05.text.upper()
        self.manager.Obs05_text = self.ids.Obs05.text.upper()


class PB06(MDScreen):
    def clear_text(self):
        self.ids.SNP06.text = ""
        self.ids.PPP06.text = ""
        self.ids.Pwd06.text = ""
        self.ids.Obs06.text = ""
        self.ids.Loid06.text = ""

    def check_box_status(self):
        if self.ids.statusok.active:
            self.StsP06 = 'OK'
            print(f"STATUS:{self.StsP06}")
        elif self.ids.statusbko.active:
            self.StsP06 = '#'
            print(f"STATUS:{self.StsP06}")
        else:
            self.StsP06 = '*'
            print(f"STATUS:{self.StsP06}")

    def Salvar_Info_PB06(self):
        self.manager.current = "tela1x16"
        self.manager.transition.direction = 'left'
        self.check_box_status()

        self.manager.StsP06_text = self.StsP06.upper()
        self.manager.Loid06_text = self.ids.Loid06.text.upper()
        self.manager.SNP06_text = self.ids.SNP06.text.upper()
        self.manager.PPP06_text = self.ids.PPP06.text.upper()
        self.manager.Pwd06_text = self.ids.Pwd06.text.upper()
        self.manager.Obs06_text = self.ids.Obs06.text.upper()


class PB07(MDScreen):
    def clear_text(self):
        self.ids.SNP07.text = ""
        self.ids.PPP07.text = ""
        self.ids.Pwd07.text = ""
        self.ids.Obs07.text = ""
        self.ids.Loid07.text = ""

    def check_box_status(self):
        if self.ids.statusok.active:
            self.StsP07 = 'OK'
            print(f"STATUS:{self.StsP07}")
        elif self.ids.statusbko.active:
            self.StsP07 = '#'
            print(f"STATUS:{self.StsP07}")
        else:
            self.StsP07 = '*'
            print(f"STATUS:{self.StsP07}")

    def Salvar_Info_PB07(self):
        self.manager.current = "tela1x16"
        self.manager.transition.direction = 'left'
        self.check_box_status()

        self.manager.StsP07_text = self.StsP07.upper()
        self.manager.Loid07_text = self.ids.Loid07.text.upper()
        self.manager.SNP07_text = self.ids.SNP07.text.upper()
        self.manager.PPP07_text = self.ids.PPP07.text.upper()
        self.manager.Pwd07_text = self.ids.Pwd07.text.upper()
        self.manager.Obs07_text = self.ids.Obs07.text.upper()


class PB08(MDScreen):
    def clear_text(self):
        self.ids.SNP08.text = ""
        self.ids.PPP08.text = ""
        self.ids.Pwd08.text = ""
        self.ids.Obs08.text = ""
        self.ids.Loid08.text = ""

    def check_box_status(self):
        if self.ids.statusok.active:
            self.StsP08 = 'OK'
            print(f"STATUS:{self.StsP08}")
        elif self.ids.statusbko.active:
            self.StsP08 = '#'
            print(f"STATUS:{self.StsP08}")
        else:
            self.StsP08 = '*'
            print(f"STATUS:{self.StsP08}")

    def Salvar_Info_PB08(self):
        self.manager.current = "tela1x16"
        self.manager.transition.direction = 'left'
        self.check_box_status()

        self.manager.StsP08_text = self.StsP08.upper()
        self.manager.Loid08_text = self.ids.Loid08.text.upper()
        self.manager.SNP08_text = self.ids.SNP08.text.upper()
        self.manager.PPP08_text = self.ids.PPP08.text.upper()
        self.manager.Pwd08_text = self.ids.Pwd08.text.upper()
        self.manager.Obs08_text = self.ids.Obs08.text.upper()


class PB09(MDScreen):
    def clear_text(self):
        self.ids.SNP09.text = ""
        self.ids.PPP09.text = ""
        self.ids.Pwd09.text = ""
        self.ids.Obs09.text = ""
        self.ids.Loid09.text = ""

    def check_box_status(self):
        if self.ids.statusok.active:
            self.StsP09 = 'OK'
            print(f"STATUS:{self.StsP09}")
        elif self.ids.statusbko.active:
            self.StsP09 = '#'
            print(f"STATUS:{self.StsP09}")
        else:
            self.StsP09 = '*'
            print(f"STATUS:{self.StsP09}")

    def Salvar_Info_PB09(self):
        self.manager.current = "tela1x16"
        self.manager.transition.direction = 'left'
        self.check_box_status()

        self.manager.StsP09_text = self.StsP09.upper()
        self.manager.Loid09_text = self.ids.Loid09.text.upper()
        self.manager.SNP09_text = self.ids.SNP09.text.upper()
        self.manager.PPP09_text = self.ids.PPP09.text.upper()
        self.manager.Pwd09_text = self.ids.Pwd09.text.upper()
        self.manager.Obs09_text = self.ids.Obs09.text.upper()


class PB10(MDScreen):
    def clear_text(self):
        self.ids.SNP10.text = ""
        self.ids.PPP10.text = ""
        self.ids.Pwd10.text = ""
        self.ids.Obs10.text = ""
        self.ids.Loid10.text = ""

    def check_box_status(self):
        if self.ids.statusok.active:
            self.StsP10 = 'OK'
            print(f"STATUS:{self.StsP10}")
        elif self.ids.statusbko.active:
            self.StsP10 = '#'
            print(f"STATUS:{self.StsP10}")
        else:
            self.StsP10 = '*'
            print(f"STATUS:{self.StsP10}")

    def Salvar_Info_PB10(self):
        self.manager.current = "tela1x16"
        self.manager.transition.direction = 'left'
        self.check_box_status()

        self.manager.StsP10_text = self.StsP10.upper()
        self.manager.Loid10_text = self.ids.Loid10.text.upper()
        self.manager.SNP10_text = self.ids.SNP10.text.upper()
        self.manager.PPP10_text = self.ids.PPP10.text.upper()
        self.manager.Pwd10_text = self.ids.Pwd10.text.upper()
        self.manager.Obs10_text = self.ids.Obs10.text.upper()


class PB11(MDScreen):
    def clear_text(self):
        self.ids.SNP11.text = ""
        self.ids.PPP11.text = ""
        self.ids.Pwd11.text = ""
        self.ids.Obs11.text = ""
        self.ids.Loid11.text = ""

    def check_box_status(self):
        if self.ids.statusok.active:
            self.StsP11 = 'OK'
            print(f"STATUS:{self.StsP11}")
        elif self.ids.statusbko.active:
            self.StsP11 = '#'
            print(f"STATUS:{self.StsP11}")
        else:
            self.StsP11 = '*'
            print(f"STATUS:{self.StsP11}")

    def Salvar_Info_PB11(self):
        self.manager.current = "tela1x16"
        self.manager.transition.direction = 'left'
        self.check_box_status()

        self.manager.StsP11_text = self.StsP11.upper()
        self.manager.Loid11_text = self.ids.Loid11.text.upper()
        self.manager.SNP11_text = self.ids.SNP11.text.upper()
        self.manager.PPP11_text = self.ids.PPP11.text.upper()
        self.manager.Pwd11_text = self.ids.Pwd11.text.upper()
        self.manager.Obs11_text = self.ids.Obs11.text.upper()


class PB12(MDScreen):
    def clear_text(self):
        self.ids.SNP12.text = ""
        self.ids.PPP12.text = ""
        self.ids.Pwd12.text = ""
        self.ids.Obs12.text = ""
        self.ids.Loid12.text = ""

    def check_box_status(self):
        if self.ids.statusok.active:
            self.StsP12 = 'OK'
            print(f"STATUS:{self.StsP12}")
        elif self.ids.statusbko.active:
            self.StsP12 = '#'
            print(f"STATUS:{self.StsP12}")
        else:
            self.StsP12 = '*'
            print(f"STATUS:{self.StsP12}")

    def Salvar_Info_PB12(self):
        self.manager.current = "tela1x16"
        self.manager.transition.direction = 'left'
        self.check_box_status()

        self.manager.StsP12_text = self.StsP12.upper()
        self.manager.Loid12_text = self.ids.Loid12.text.upper()
        self.manager.SNP12_text = self.ids.SNP12.text.upper()
        self.manager.PPP12_text = self.ids.PPP12.text.upper()
        self.manager.Pwd12_text = self.ids.Pwd12.text.upper()
        self.manager.Obs12_text = self.ids.Obs12.text.upper()


class PB13(MDScreen):
    def clear_text(self):
        self.ids.SNP13.text = ""
        self.ids.PPP13.text = ""
        self.ids.Pwd13.text = ""
        self.ids.Obs13.text = ""
        self.ids.Loid13.text = ""
    def check_box_status(self):
        if self.ids.statusok.active:
            self.StsP13 = 'OK'
            print(f"STATUS:{self.StsP13}")
        elif self.ids.statusbko.active:
            self.StsP13 = '#'
            print(f"STATUS:{self.StsP13}")
        else:
            self.StsP13 = '*'
            print(f"STATUS:{self.StsP13}")

    def Salvar_Info_PB13(self):
        self.manager.current = "tela1x16"
        self.manager.transition.direction = 'left'
        self.check_box_status()

        self.manager.StsP13_text = self.StsP13.upper()
        self.manager.Loid13_text = self.ids.Loid13.text.upper()
        self.manager.SNP13_text = self.ids.SNP13.text.upper()
        self.manager.PPP13_text = self.ids.PPP13.text.upper()
        self.manager.Pwd13_text = self.ids.Pwd13.text.upper()
        self.manager.Obs13_text = self.ids.Obs13.text.upper()


class PB14(MDScreen):
    def clear_text(self):
        self.ids.SNP14.text = ""
        self.ids.PPP14.text = ""
        self.ids.Pwd14.text = ""
        self.ids.Obs14.text = ""
        self.ids.Loid14.text = ""
    def check_box_status(self):
        if self.ids.statusok.active:
            self.StsP14 = 'OK'
            print(f"STATUS:{self.StsP14}")
        elif self.ids.statusbko.active:
            self.StsP14 = '#'
            print(f"STATUS:{self.StsP14}")
        else:
            self.StsP14 = '*'
            print(f"STATUS:{self.StsP14}")

    def Salvar_Info_PB14(self):
        self.manager.current = "tela1x16"
        self.manager.transition.direction = 'left'
        self.check_box_status()

        self.manager.StsP14_text = self.StsP14.upper()
        self.manager.Loid14_text = self.ids.Loid14.text.upper()
        self.manager.SNP14_text = self.ids.SNP14.text.upper()
        self.manager.PPP14_text = self.ids.PPP14.text.upper()
        self.manager.Pwd14_text = self.ids.Pwd14.text.upper()
        self.manager.Obs14_text = self.ids.Obs14.text.upper()


class PB15(MDScreen):
    def clear_text(self):
        self.ids.SNP15.text = ""
        self.ids.PPP15.text = ""
        self.ids.Pwd15.text = ""
        self.ids.Obs15.text = ""
        self.ids.Loid15.text = ""
    def check_box_status(self):
        if self.ids.statusok.active:
            self.StsP15 = 'OK'
            print(f"STATUS:{self.StsP15}")
        elif self.ids.statusbko.active:
            self.StsP15 = '#'
            print(f"STATUS:{self.StsP15}")
        else:
            self.StsP15 = '*'
            print(f"STATUS:{self.StsP15}")

    def Salvar_Info_PB15(self):
        self.manager.current = "tela1x16"
        self.manager.transition.direction = 'left'
        self.check_box_status()

        self.manager.StsP15_text = self.StsP15.upper()
        self.manager.Loid15_text = self.ids.Loid15.text.upper()
        self.manager.SNP15_text = self.ids.SNP15.text.upper()
        self.manager.PPP15_text = self.ids.PPP15.text.upper()
        self.manager.Pwd15_text = self.ids.Pwd15.text.upper()
        self.manager.Obs15_text = self.ids.Obs15.text.upper()


class PB16(MDScreen):
    def clear_text(self):
        self.ids.SNP16.text = ""
        self.ids.PPP16.text = ""
        self.ids.Pwd16.text = ""
        self.ids.Obs16.text = ""
        self.ids.Loid16.text = ""

    def check_box_status(self):
        if self.ids.statusok.active:
            self.StsP16 = 'OK'
            print(f"STATUS:{self.StsP16}")
        elif self.ids.statusbko.active:
            self.StsP16 = '#'
            print(f"STATUS:{self.StsP16}")
        else:
            self.StsP16 = '*'
            print(f"STATUS:{self.StsP16}")

    def Salvar_Info_PB16(self):
        self.manager.current = "tela1x16"
        self.manager.transition.direction = 'left'
        self.check_box_status()

        self.manager.StsP16_text = self.StsP16.upper()
        self.manager.Loid16_text = self.ids.Loid16.text.upper()
        self.manager.SNP16_text = self.ids.SNP16.text.upper()
        self.manager.PPP16_text = self.ids.PPP16.text.upper()
        self.manager.Pwd16_text = self.ids.Pwd16.text.upper()
        self.manager.Obs16_text = self.ids.Obs16.text.upper()
